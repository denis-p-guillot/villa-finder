#!/usr/bin/env python3
"""
Unified villa scraper: all sources into one CSV format.
Usage: --source all | rumah123 | villa-bali | balilongterm | balicoconut
Output: single CSV (villa_listings.csv) with unified columns; only rows with title and price_idr.
"""

import argparse
import csv
import json
import queue
import random
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urljoin, urlparse, unquote, parse_qs, urlencode, urlunparse

import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# Proxy rotation: each request uses the next proxy in round-robin order
# ---------------------------------------------------------------------------

PROXY_SPECS = [
    {"host": "31.59.20.176", "port": 6754},
    {"host": "23.95.150.145", "port": 6114},
    {"host": "198.23.239.134", "port": 6540},
    {"host": "45.38.107.97", "port": 6014},
    {"host": "107.172.163.27", "port": 6543},
    {"host": "198.105.121.200", "port": 6462},
    {"host": "64.137.96.74", "port": 6641},
    {"host": "216.10.27.159", "port": 6837},
    {"host": "142.111.67.146", "port": 5611},
    {"host": "191.96.254.138", "port": 6185},
]
PROXY_USER = "nnfetybf"
PROXY_PASS = "j0y8vamctfuz"

_proxy_lock = threading.Lock()
_proxy_index = 0
_proxyless_mode = False


def _next_proxy_for_requests() -> tuple[dict | None, str]:
    """Return (proxies dict or None for no proxy, ip_display). If proxyless mode, returns (None, 'proxyless')."""
    global _proxy_index
    with _proxy_lock:
        if _proxyless_mode:
            return None, "proxyless"
        spec = PROXY_SPECS[_proxy_index % len(PROXY_SPECS)]
        _proxy_index += 1
    ip_display = f"{spec['host']}:{spec['port']}"
    url = f"http://{PROXY_USER}:{PROXY_PASS}@{spec['host']}:{spec['port']}/"
    return {"http": url, "https": url}, ip_display


def _next_proxy_for_playwright() -> tuple[dict | None, str]:
    """Return (proxy dict or None for no proxy, ip_display). If proxyless mode, returns (None, 'proxyless')."""
    global _proxy_index
    with _proxy_lock:
        if _proxyless_mode:
            return None, "proxyless"
        spec = PROXY_SPECS[_proxy_index % len(PROXY_SPECS)]
        _proxy_index += 1
    ip_display = f"{spec['host']}:{spec['port']}"
    proxy = {
        "server": f"http://{spec['host']}:{spec['port']}",
        "username": PROXY_USER,
        "password": PROXY_PASS,
    }
    return proxy, ip_display


from villa_csv import MAIN_CSV_COLUMNS, get_logger, setup_logging

log = get_logger()


def _is_proxy_error(e: BaseException) -> bool:
    """True if the exception is likely due to proxy/connection failure."""
    err_str = str(e).lower()
    if "proxy" in err_str or "err_proxy" in err_str or "connection" in err_str or "timeout" in err_str:
        return True
    exc_name = type(e).__name__
    if exc_name in ("ProxyError", "ConnectTimeout", "ConnectionError", "ReadTimeout", "ConnectError"):
        return True
    return False


def _switch_to_proxyless() -> None:
    """Set global proxyless mode and log once. Subsequent requests will not use proxies."""
    global _proxyless_mode
    with _proxy_lock:
        if _proxyless_mode:
            return
        _proxyless_mode = True
    log.warning(
        "Proxies stopped responding; switching to proxyless mode. Scraping will continue without proxies."
    )

try:
    from scrape_balilongterm import run_balilongterm
except ImportError:
    run_balilongterm = None
try:
    from scrape_balicoconut import run_balicoconut
except ImportError:
    run_balicoconut = None
try:
    from scrape_balihomeimmo import run_balihomeimmo
except ImportError:
    run_balihomeimmo = None
try:
    from scrape_balirealty import run_balirealty
except ImportError:
    run_balirealty = None

# ---------------------------------------------------------------------------
# Shared CSV format (canonical columns from villa_csv)
# ---------------------------------------------------------------------------


def _csv_cell(s: str) -> str:
    """Sanitize a value for CSV: no commas, no newlines, no double-quotes."""
    if not isinstance(s, str):
        return ""
    if "push(" in s or '"formats"' in s or '"locale"' in s:
        return ""
    s = str(s).replace('"', "'").replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = re.sub(r" +", " ", s).strip()
    if len(s) > 8000:
        s = s[:8000] + "..."
    return s.replace(",", "")


# Rumah123: parse "Rp X /hari" -> (price_idr, duration)
DURATION_MAP = {
    "/hari": "day", "per hari": "day", "/minggu": "week", "per minggu": "week",
    "/bulan": "month", "per bulan": "month", "/tahun": "year", "per tahun": "year",
    "/year": "year", "/month": "month", "/day": "day",
}

RUMAH123_BALI_LOCATIONS = [
    "Abang",
    "Abiansemal",
    "Banjar",
    "Banjarangkan",
    "Bangli",
    "Baturiti",
    "Bebandem",
    "Blahbatuh",
    "Buleleng",
    "Busungbiu",
    "Dawan",
    "Denpasar Barat",
    "Denpasar Selatan",
    "Denpasar Timur",
    "Denpasar Utara",
    "Gerokgak",
    "Gianyar",
    "Jembrana",
    "Karangasem",
    "Kediri",
    "Kerambitan",
    "Kintamani",
    "Klungkung",
    "Kubu",
    "Kubutambahan",
    "Kuta",
    "Kuta Selatan",
    "Kuta Utara",
    "Manggis",
    "Marga",
    "Melaya",
    "Mendoyo",
    "Mengwi",
    "Negara",
    "Nusa Penida",
    "Payangan",
    "Pekutatan",
    "Penebel",
    "Petang",
    "Pupuan",
    "Rendang",
    "Sawan",
    "Selemadeg",
    "Selemadeg Barat",
    "Selemadeg Timur",
    "Selat",
    "Seririt",
    "Sidemen",
    "Sukawati",
    "Sukasada",
    "Susut",
    "Tabanan",
    "Tampaksiring",
    "Tegallalang",
    "Tejakula",
    "Tembuku",
    "Ubud",
]


def _parse_price_idr_rumah123(price_text: str) -> tuple[str, str]:
    """
    Normalize various Indonesian "Rp ..." price formats into a canonical
    "Rp X" string and extract rental duration (day/month/year).

    This now also supports plain absolute IDR formats like
    "Rp 10.550.000.000" (as used by OLX) in addition to Rumah123-style
    textual amounts such as "Rp 10 miliar", "Rp 15 juta", etc.
    """
    price_text = (price_text or "").strip()
    duration = ""

    # Extract and strip duration markers such as "/bulan", "per tahun", etc.
    lower_full = price_text.lower()
    for key, eng in DURATION_MAP.items():
        if key in lower_full:
            duration = eng
            # Keep only the part before the duration marker for numeric parsing.
            price_text = lower_full.split(key)[0].strip()
            break

    if not price_text:
        return "", duration

    # Core numeric extraction:
    # - Accept both "10.550.000.000" and "10,550,000,000" and any mix of
    #   non-digit separators.
    digits_only = re.sub(r"[^\d]", "", price_text)
    if not digits_only:
        return "", duration

    try:
        amount = int(digits_only)
    except ValueError:
        return "", duration

    lower = price_text.lower()

    # Rumah123 textual magnitudes:
    if "miliar" in lower or "milliar" in lower:
        amount *= 1_000_000_000
    elif "juta" in lower:
        amount *= 1_000_000
    elif "ribu" in lower:
        amount *= 1_000
    else:
        # Heuristic for bare numbers without magnitude words:
        # - If the numeric value is very small, treat it as "juta".
        # - If it's already large (>= 10,000) assume it's the full IDR amount.
        if amount < 10_000:
            amount *= 1_000_000

    amount_int = int(amount)
    return f"Rp {amount_int:,}".replace(",", "."), duration


def flatten_row(row: dict, source: str = "rumah123") -> dict:
    """Normalize row to MAIN_CSV_COLUMNS. Handles Rumah123 (price + ID keys) and Villa-Bali (price_idr + direct keys)."""
    raw = {}
    for k, v in row.items():
        if isinstance(v, (list, dict)):
            raw[k] = _csv_cell(json.dumps(v, ensure_ascii=False)) if v else ""
        else:
            raw[k] = _csv_cell(str(v or ""))

    if source == "rumah123":
        facilities_parts = [
            raw.get("facilities_rumah") or "",
            raw.get("facilities_perumahan") or "",
            raw.get("perabotan") or "",
        ]
        facilities_str = " | ".join(p for p in facilities_parts if p).strip() or ""
        price_raw = raw.get("price", "")
        price_idr, duration = _parse_price_idr_rumah123(price_raw)
        out = {
            "title": (raw.get("title", "") or "").replace(",", ""),
            "price_idr": price_idr,
            "duration": duration,
            "location": raw.get("location", ""),
            "bedrooms": raw.get("Kamar_Tidur", ""),
            "bathrooms": raw.get("Kamar_Mandi", ""),
            "land_size_m2": raw.get("Luas_Tanah", ""),
            "building_size_m2": raw.get("Luas_Bangunan", ""),
            "certificate": raw.get("Sertifikat", ""),
            "description": (raw.get("description", "") or "").replace(",", ""),
            "facilities": facilities_str.replace(",", ""),
            "agent_name": (raw.get("agent_name", "") or "").replace(",", ""),
            "updated_by": raw.get("updated_by", ""),
            "main_image": raw.get("main_image", ""),
            "image_urls": raw.get("image_urls", ""),
            "url": raw.get("url", ""),
        }
    else:
        # villa-bali: already has price_idr, duration, etc.
        out = {
            "title": (raw.get("title", "") or "").replace(",", ""),
            "price_idr": raw.get("price_idr", ""),
            "duration": raw.get("duration", "day"),
            "location": raw.get("location", ""),
            "bedrooms": raw.get("bedrooms", ""),
            "bathrooms": raw.get("bathrooms", ""),
            "land_size_m2": raw.get("land_size_m2", ""),
            "building_size_m2": raw.get("building_size_m2", ""),
            "certificate": raw.get("certificate", ""),
            "description": (raw.get("description", "") or "").replace(",", ""),
            "facilities": (raw.get("facilities", "") or "").replace(",", ""),
            "agent_name": (raw.get("agent_name", "") or "").replace(",", ""),
            "updated_by": raw.get("updated_by", ""),
            "main_image": raw.get("main_image", ""),
            "url": raw.get("url", ""),
        }
    return {k: _csv_cell(str(v)) for k, v in out.items()}


def _row_has_title_and_price(row: dict, source: str) -> bool:
    """True if row has both title and price_idr."""
    if not row:
        return False
    title_ok = bool((row.get("title") or "").strip())
    if source == "rumah123":
        price_idr, _ = _parse_price_idr_rumah123(row.get("price") or "")
    else:
        price_idr = (row.get("price_idr") or "").strip()
    return title_ok and bool(price_idr)


def _to_int_or_none_from_idr(price_idr: str) -> int | None:
    """Convert 'Rp 30.000.000' → 30000000 (or None on failure)."""
    if not price_idr:
        return None
    digits = re.sub(r"[^\d]", "", str(price_idr))
    try:
        return int(digits) if digits else None
    except ValueError:
        return None


def _row_to_portal_villa(row: dict, source: str) -> dict:
    """
    Convert a unified CSV row into the JSON villa schema requested by the user.

    This uses whatever fields are available in the flattened row and applies
    sensible defaults / derivations for the missing ones.
    """
    price_idr_str = row.get("price_idr") or ""
    price_idr = _to_int_or_none_from_idr(price_idr_str) or 0

    duration = (row.get("duration") or "").strip().lower() or "day"
    if duration in ("bulan", "per bulan", "/bulan"):
        duration = "month"
    elif duration in ("tahun", "per tahun", "/tahun"):
        duration = "year"

    if duration == "month":
        rental_type = "monthly"
        price_monthly = price_idr
        price_yearly = price_idr * 12 if price_idr else 0
    elif duration == "year":
        rental_type = "yearly"
        price_yearly = price_idr
        price_monthly = round(price_idr / 12) if price_idr else 0
    else:
        rental_type = "daily"
        price_monthly = price_idr * 30 if price_idr else 0
        price_yearly = price_idr * 365 if price_idr else 0

    location = (row.get("location") or "").strip()
    title = (row.get("title") or "").strip()

    try:
        bedrooms = int(row.get("bedrooms") or 0)
    except (TypeError, ValueError):
        bedrooms = 0
    try:
        bathrooms = int(row.get("bathrooms") or 0)
    except (TypeError, ValueError):
        bathrooms = 0

    def _to_float_or_zero(v):
        try:
            return float(v) if v not in (None, "") else 0
        except (TypeError, ValueError):
            return 0

    land_size = _to_float_or_zero(row.get("land_size_m2"))
    building_size = _to_float_or_zero(row.get("building_size_m2"))

    description = (row.get("description") or "").strip()
    long_description = description  # we do not have a separate long description

    facilities_raw = (row.get("facilities") or "").strip()
    facilities_list = [f.strip() for f in facilities_raw.split("|") if f.strip()] if facilities_raw else []

    # We don't have a clean split between facilities/amenities/features from this source,
    # so we use facilities for all three buckets and keep the raw string as-is.
    amenities_list = list(facilities_list)
    features_list = []

    main_image_url = (row.get("main_image") or "").strip()

    villa = {
        "source_url": row.get("url") or "",
        "detail_url": row.get("url") or "",
        "title": title,
        "location_name": location,
        "location": location,
        "area_name": location,
        "area": location,
        "managed_by": "agent",
        "rental_type": rental_type,
        "contact_email": "",  # not available in current scraper
        "price_idr": price_idr,
        "price": price_idr,
        "price_monthly": int(price_monthly),
        "price_yearly": int(price_yearly),
        "minimum_stay": "1 " + ("month" if rental_type == "monthly" else "year" if rental_type == "yearly" else "day"),
        "bedrooms": bedrooms,
        "beds": bedrooms,
        "bathrooms": bathrooms,
        "baths": bathrooms,
        "land_size_m2": land_size,
        "land_size": land_size,
        "building_size_m2": building_size,
        "building_size": building_size,
        "description": description,
        "long_description": long_description,
        "facilities": facilities_list,
        "amenities": amenities_list,
        "features": features_list,
        "facilities_raw": facilities_raw,
        "main_image_url": main_image_url,
        "image_urls": [],
        "images": [],
        "image_urls_raw": "",
    }
    return villa


# ---------------------------------------------------------------------------
# User-Agent rotation (new UA for every request)
# ---------------------------------------------------------------------------

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:122.0) Gecko/20100101 Firefox/122.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
]


def _random_user_agent() -> str:
    return random.choice(USER_AGENTS)


def _rumah123_headers() -> dict:
    return {
        "User-Agent": _random_user_agent(),
        "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
    }


# ---------------------------------------------------------------------------
# Rumah123
# ---------------------------------------------------------------------------

# When security verification is detected in headless, wait this many seconds before
# giving up and retrying with a visible browser (so the real browser kicks in quickly).
RUMAH123_SECURITY_RETRY_SEC = 15

# Target zone for the verification checkbox only (viewport 1920x1080).
# Zone shifted so center is ~20px to the right of the checkbox center (per calibration).
RUMAH123_VERIFICATION_CLICK_X_MIN = 520
RUMAH123_VERIFICATION_CLICK_X_MAX = 600
RUMAH123_VERIFICATION_CLICK_Y_MIN = 310
RUMAH123_VERIFICATION_CLICK_Y_MAX = 350
RUMAH123_VERIFICATION_CLICK_STEP = 6
RUMAH123_VERIFICATION_CLICK_MAX = 30  # max clicks per wave inside the zone

# Page text that indicates the "verify you are human" challenge.
RUMAH123_VERIFICATION_PHRASES = (
    "melakukan verifikasi keamanan",
    "verifikasi bahwa anda adalah manusia",
    "buktikan bahwa anda adalah manusia",
)


def _rumah123_page_has_verification(html_lower: str) -> bool:
    """True if the page shows the human verification challenge (checkbox must be clicked in browser)."""
    return any(phrase in html_lower for phrase in RUMAH123_VERIFICATION_PHRASES)


RUMAH123_BASE = "https://www.rumah123.com"
RUMAH123_LISTING = f"{RUMAH123_BASE}/sewa/bali/villa/"


def _rumah123_session():
    s = requests.Session()
    # Do not set fixed User-Agent here; use _rumah123_headers() on every request
    return s


def _rumah123_extract_max_pages_from_html(html: str) -> int | None:
    """
    Inspect listing HTML and infer the maximum page number from pagination links.
    We look for links like '?page=2', '?page=10', etc. and take the highest value.
    Returns None if no pagination links are found.
    """
    if not html:
        return None
    try:
        nums = [int(m) for m in re.findall(r"[?&]page=(\d+)", html)]
        return max(nums) if nums else None
    except Exception:
        return None


def _find_text_in_body(soup: BeautifulSoup, pattern):
    for el in soup.find_all(string=pattern):
        if not el.parent:
            continue
        p = el.parent
        while p:
            if getattr(p, "name", None) in ("script", "style", "noscript"):
                break
            p = p.parent
        else:
            return el
    return None


def _rumah123_normalize_image_url(url: str) -> str:
    """
    Normalize Rumah123 image URLs:
    - Ensure absolute https:// URL
    - Prefer 1080x720-fit variant when a size segment is present.
    """
    url = (url or "").strip()
    if not url:
        return ""
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    # Upgrade any size segment (e.g. 720x420-crop, 360x401-inside, 250x160-fit) to 1080x720-fit.
    url = re.sub(r"/\d+x\d+-(?:crop|fit|inside)/", "/1080x720-fit/", url)
    return url


def _rumah123_extract_district_from_json(html: str) -> str:
    """
    Extract the district name from the embedded JSON config in the detail page, if present.
    Example fragment:
      "location":{"uuid":"...","level":3,"name":"Jimbaran",...,
                  "district":{"id":"...","name":"Jimbaran",...}, ...}
    """
    if not html:
        return ""
    try:
        m = re.search(r'"district"\s*:\s*\{[^}]*"name"\s*:\s*"([^"]+)"', html)
        if m:
            return m.group(1).strip()
    except re.error:
        return ""
    return ""


def _rumah123_normalize_location(raw_location: str) -> str:
    """
    Given a free-text location like 'Jimbaran, Badung' or 'Villa in Ubud Bali',
    normalize it to one of the known Bali districts if possible.
    Returns the matched canonical name, or the original string if no match.
    """
    loc = (raw_location or "").strip()
    if not loc:
        return ""
    lower = loc.lower()
    # Prefer longer names first (e.g. 'Kuta Selatan' before 'Kuta').
    for name in sorted(RUMAH123_BALI_LOCATIONS, key=len, reverse=True):
        if name.lower() in lower:
            return name
    return loc


def _rumah123_extract_images_from_html(html: str, max_images: int = 40) -> list[str]:
    """
    Extract a list of image URLs from raw Rumah123 detail HTML.

    Preference order:
    1) 1080x720-fit images (highest quality slider images present in JSON).
    2) Slider/thumbnail sizes (e.g. 720x420-crop, 360x401-inside, 250x160-fit).
    3) Any other pic.rumah123.com / picture.rumah123.com JPG/PNG/WEBP.
    """
    if not html:
        return []
    try:
        candidates = re.findall(
            r"https://(?:pic|picture)\.rumah123\.com/[^\s\"'>)]+",
            html,
        )
    except re.error:
        return []

    if not candidates:
        return []

    # Deduplicate while preserving original order.
    seen: set[str] = set()
    unique: list[str] = []
    for u in candidates:
        if u not in seen:
            seen.add(u)
            unique.append(u)

    hi_res: list[str] = []
    slider_sizes: list[str] = []
    others: list[str] = []

    for u in unique:
        ul = u.lower()
        if not ul.endswith((".jpg", ".jpeg", ".png", ".webp")):
            continue
        if "/1080x720-fit/" in ul:
            hi_res.append(u)
        elif any(s in ul for s in ("/720x420-crop/", "/360x401-inside/", "/250x160-fit/")):
            slider_sizes.append(u)
        else:
            others.append(u)

    ordered = hi_res + slider_sizes + others
    normalized: list[str] = []
    for u in ordered:
        normalized.append(_rumah123_normalize_image_url(u))
        if len(normalized) >= max_images:
            break
    return normalized


def _rumah123_extract_main_image_from_html(html: str) -> str:
    """
    Backwards-compatible helper: return only the first/best image URL.
    """
    images = _rumah123_extract_images_from_html(html, max_images=1)
    if not images:
        return ""
    return images[0]


def rumah123_fetch_listing_links(session: requests.Session, page: int = 1) -> tuple[list[str], bool]:
    """Fetch one listing page. Returns (links, ok). ok=False on network/429 exhaustion so caller can try next page."""
    url = RUMAH123_LISTING if page <= 1 else f"{RUMAH123_LISTING}?page={page}"
    saw_429 = False
    r = None
    for attempt in range(4):
        try:
            proxies, ip = _next_proxy_for_requests()
            log.info("rumah123 | listing page %s via proxy %s", page, ip)
            kwargs = {"timeout": 45, "headers": _rumah123_headers()}
            if proxies is not None:
                kwargs["proxies"] = proxies
            r = session.get(url, **kwargs)
            # Retry up to 3 times with backoff when rate-limited
            for retry in range(3):
                if r.status_code != 429:
                    break
                saw_429 = True
                wait = 20 + (retry * 15)  # 20s, then 35s, then 50s
                log.info("rumah123 | 429 rate limit, waiting %ss before retry %s", wait, retry + 1)
                time.sleep(wait)
                proxies, ip = _next_proxy_for_requests()
                log.info("rumah123 | listing page %s retry via proxy %s", page, ip)
                kwargs = {"timeout": 45, "headers": _rumah123_headers()}
                if proxies is not None:
                    kwargs["proxies"] = proxies
                r = session.get(url, **kwargs)
            r.raise_for_status()
            break
        except Exception as e:
            if _is_proxy_error(e):
                _switch_to_proxyless()
                try:
                    log.info("rumah123 | listing page %s retry without proxy", page)
                    r = session.get(url, timeout=45, headers=_rumah123_headers())
                    for retry in range(3):
                        if r.status_code != 429:
                            break
                        saw_429 = True
                        time.sleep(20 + retry * 15)
                        r = session.get(url, timeout=45, headers=_rumah123_headers())
                    r.raise_for_status()
                    break
                except Exception as e2:
                    if attempt < 3:
                        time.sleep(5 * (attempt + 1))
                    else:
                        log.warning("rumah123 | page %s failed after retries: %s", page, e2)
                        return ([], False)
            elif attempt < 3:
                time.sleep(5 * (attempt + 1))
            else:
                log.warning("rumah123 | page %s failed after retries: %s", page, e)
                return ([], False)
    if r is None:
        return ([], False)
    # Cooldown after 429 so next page request doesn't immediately hit rate limit again
    if saw_429:
        cooldown = 12
        log.info("rumah123 | cooldown %ss after 429 before next page", cooldown)
        time.sleep(cooldown)
    soup = BeautifulSoup(r.text, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href.startswith("/properti/") or "/agent-" in href or "/independent-property-agent/" in href:
            continue
        if re.search(r"-(?:hor|vlr)\d+/?$", href):
            path = href.rstrip("/") + "/" if not href.endswith("/") else href
            full = urljoin(RUMAH123_BASE, path)
            if full not in links:
                links.append(full)
    return (links, True)


def rumah123_parse_spec_table(soup: BeautifulSoup) -> dict:
    out = {}
    for block in soup.find_all(["section", "div"], class_=re.compile(r"spec|info|detail", re.I)):
        texts = block.get_text(separator="|", strip=True)
        if "Luas Tanah" not in texts and "Kamar Tidur" not in texts:
            continue
        parts = [p.strip() for p in texts.replace("\n", "|").split("|") if p.strip()]
        for i, p in enumerate(parts):
            if p in ("Luas Tanah", "Luas Bangunan", "Kamar Tidur", "Kamar Mandi", "Carport", "Sertifikat", "Daya Listrik", "Kondisi / Tahun Renovasi"):
                if i + 1 < len(parts):
                    val = parts[i + 1].strip()
                    if val and val != p:
                        out[p] = val
    labels = ("Luas Tanah", "Luas Bangunan", "Kamar Tidur", "Kamar Mandi", "Carport", "Sertifikat", "Daya Listrik")
    for label in labels:
        if label in out:
            continue
        el = soup.find(string=re.compile(re.escape(label), re.I))
        if el and el.parent:
            parent = el.parent
            n = parent.find_next_sibling()
            if n:
                v = n.get_text(strip=True)
                if v:
                    out[label] = v
    if not out and soup.body:
        body_text = soup.body.get_text(separator="\n", strip=True)
        for label in labels:
            if label in out:
                continue
            m = re.search(re.escape(label) + r"\s*\n\s*([^\n]{1,50})", body_text)
            if m:
                out[label] = m.group(1).strip()
    return out


def rumah123_parse_facilities(soup: BeautifulSoup) -> dict:
    out = {"Fasilitas Rumah": [], "Fasilitas Perumahan": [], "Perabotan": []}
    section = _find_text_in_body(soup, re.compile(r"Fasilitas", re.I))
    if not section:
        return out
    root = section.parent
    while root and getattr(root, "name", None) not in ("section", "div"):
        root = getattr(root, "parent", None)
    if not root:
        return out
    for h in root.find_all(["h3", "h4", "strong"], string=re.compile(r"Fasilitas Rumah|Fasilitas Perumahan|Perabotan", re.I)):
        key = h.get_text(strip=True)
        if key not in out:
            continue
        container = h.find_parent("div") or h
        next_ = container.find_next_sibling() or container.find_next()
        if next_ and getattr(next_, "find_all", None):
            items = next_.find_all(["li", "span", "div"], recursive=True)
            texts = [node.get_text(strip=True) for node in items[:50] if node.get_text(strip=True) and len(node.get_text(strip=True)) < 80]
            out[key] = list(dict.fromkeys(texts))[:30] if texts else []
    return out


def _rumah123_soup_to_row(soup: BeautifulSoup, url: str, html: str | None = None) -> dict:
    """Parse Rumah123 detail page HTML (from requests or Playwright) into a row dict."""
    h1 = soup.find("h1")
    title = h1.get_text(strip=True).replace(",", "") if h1 else ""
    price_text = ""
    el = _find_text_in_body(soup, re.compile(r"Rp\s*[\d,\.]+", re.I))
    if el:
        price_text = el.strip()
    # Location: prefer district name from embedded JSON, then fall back to HTML heuristics.
    location = ""
    if html:
        district = _rumah123_extract_district_from_json(html)
        if district:
            location = district
    if not location:
        for el in soup.find_all(
            string=re.compile(
                r"(Abang|Abiansemal|Banjar|Banjarangkan|Bangli|Baturiti|Bebandem|Blahbatuh|Buleleng|Busungbiu|Dawan|Denpasar Barat|Denpasar Selatan|Denpasar Timur|Denpasar Utara|Gerokgak|Gianyar|Jembrana|Karangasem|Kediri|Kerambitan|Kintamani|Klungkung|Kubu|Kubutambahan|Kuta|Kuta Selatan|Kuta Utara|Manggis|Marga|Melaya|Mendoyo|Mengwi|Negara|Nusa Penida|Payangan|Pekutatan|Penebel|Petang|Pupuan|Rendang|Sawan|Selemadeg(?: Barat| Timur)?|Selat|Seririt|Sidemen|Sukawati|Sukasada|Susut|Tabanan|Tampaksiring|Tegallalang|Tejakula|Tembuku|Ubud)",
                re.I,
            )
        ):
            if el.parent and getattr(el.parent, "name", None) in ("script", "style", "noscript"):
                continue
            t = el.strip()
            tl = t.lower()
            if tl.startswith("http://") or tl.startswith("https://") or "://" in tl:
                continue
            if 3 <= len(t) <= 120:
                location = t
                break
    if not location and h1:
        p = h1.find_next(["p", "div", "span"])
        if p:
            location = p.get_text(strip=True).split(",")[0].strip()[:80]
    desc = ""
    d = _find_text_in_body(soup, re.compile(r"Deskripsi", re.I))
    if d and d.parent:
        parent = d.parent
        while parent and parent.name not in ("section", "div"):
            parent = parent.parent
        if parent:
            for s in parent.find_all(["p", "div"]):
                t = s.get_text(strip=True)
                if len(t) > 50 and "Deskripsi" not in t and "push(" not in t:
                    desc = t[:5000]
                    break
    # Normalize location to a canonical Bali district name when possible.
    location = _rumah123_normalize_location(location)

    updated_by = ""
    agent_name = ""
    for el in soup.find_all(string=re.compile(r"Diperbarui.*oleh", re.I)):
        if el.parent and getattr(el.parent, "name", None) not in ("script", "style"):
            updated_by = el.strip()
            break
    agent_el = soup.find("a", href=re.compile(r"agen-properti|independent-property-agent"))
    if agent_el:
        agent_name = agent_el.get_text(strip=True)
    specs = rumah123_parse_spec_table(soup)
    facilities = rumah123_parse_facilities(soup)

    # Images: collect as many slider/gallery URLs as possible from the raw HTML.
    image_list: list[str] = _rumah123_extract_images_from_html(html, max_images=40) if html else []

    # Main image: prefer the first from the collected list, then fall back to og:image / first <img>.
    main_image = image_list[0] if image_list else ""
    og = soup.find("meta", property="og:image")
    if not main_image and og and og.get("content") and "picture.rumah123.com" in (og.get("content") or ""):
        main_image = (og["content"] or "").strip()
        if "portal-api/image/og" in main_image:
            m = re.search(r"src=([^&\s]+)", main_image)
            if m:
                main_image = unquote(m.group(1).strip())
                if not main_image.startswith("http"):
                    main_image = "https://" + main_image
    if not main_image:
        for img in soup.find_all("img", src=True):
            src = (img.get("src") or "").strip()
            if "picture.rumah123.com" in src and ".jpg" in src:
                main_image = src if src.startswith("http") else "https://" + src.lstrip("/")
                break

    main_image = _rumah123_normalize_image_url(main_image)
    if not image_list and main_image:
        image_list = [main_image]

    return {
        "url": url,
        "main_image": main_image,
        "image_urls": " | ".join(image_list),
        "title": title,
        "price": price_text,
        "location": location,
        "updated_by": updated_by,
        "agent_name": agent_name,
        "description": desc,
        **{k.replace(" ", "_").replace("/", "_"): v for k, v in specs.items()},
        "facilities_rumah": " | ".join(facilities["Fasilitas Rumah"]) if isinstance(facilities["Fasilitas Rumah"], list) else "",
        "facilities_perumahan": " | ".join(facilities["Fasilitas Perumahan"]) if isinstance(facilities["Fasilitas Perumahan"], list) else "",
        "perabotan": " | ".join(facilities["Perabotan"]) if isinstance(facilities["Perabotan"], list) else "",
    }


def rumah123_parse_detail(session: requests.Session, url: str, delay: float = 1.0) -> dict | None:
    """Fetch detail page via requests and return parsed row (used when Playwright not available)."""
    time.sleep(delay)
    for attempt in range(3):
        try:
            proxies, ip = _next_proxy_for_requests()
            log.info("rumah123 | detail page %s via proxy %s", url, ip)
            kwargs = {"timeout": 45, "headers": _rumah123_headers()}
            if proxies is not None:
                kwargs["proxies"] = proxies
            r = session.get(url, **kwargs)
            for _ in range(2):
                if r.status_code != 429:
                    break
                time.sleep(20)
                proxies, ip = _next_proxy_for_requests()
                log.info("rumah123 | detail page %s retry via proxy %s", url, ip)
                kwargs = {"timeout": 45, "headers": _rumah123_headers()}
                if proxies is not None:
                    kwargs["proxies"] = proxies
                r = session.get(url, **kwargs)
            r.raise_for_status()
            html = r.text
            soup = BeautifulSoup(html, "html.parser")
            return _rumah123_soup_to_row(soup, url, html=html)
        except Exception as e:
            if _is_proxy_error(e):
                _switch_to_proxyless()
                try:
                    log.info("rumah123 | detail page %s retry without proxy", url[:80])
                    r = session.get(url, timeout=45, headers=_rumah123_headers())
                    for _ in range(2):
                        if r.status_code != 429:
                            break
                        time.sleep(20)
                        r = session.get(url, timeout=45, headers=_rumah123_headers())
                    r.raise_for_status()
                    html = r.text
                    soup = BeautifulSoup(html, "html.parser")
                    return _rumah123_soup_to_row(soup, url, html=html)
                except Exception as e2:
                    log.debug("rumah123 | error %s: %s", url[:60], e2)
                    return None
            if attempt < 2:
                time.sleep(4 * (attempt + 1))
            else:
                log.debug("rumah123 | error %s: %s", url[:60], e)
                return None
    return None


def _rumah123_browser(headless: bool = False):
    """Launch Chromium for Rumah123 (visible by default). Uses next proxy in rotation. Returns (page, browser, playwright, proxy_display)."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None
    pw = sync_playwright().start()
    browser = pw.chromium.launch(
        headless=headless,
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
    )
    proxy, ip_display = _next_proxy_for_playwright()
    log.info("rumah123 | browser using %s", ip_display)
    context_options = {
        "user_agent": _random_user_agent(),
        "locale": "id-ID",
        "viewport": {"width": 1920, "height": 1080},
    }
    if proxy is not None:
        context_options["proxy"] = proxy
    context = browser.new_context(**context_options)
    # Do not load image contents to reduce bandwidth and speed up scraping.
    try:
        context.route(
            "**/*",
            lambda route, request: route.abort()
            if request.resource_type == "image"
            else route.continue_(),
        )
    except Exception:
        # If routing fails for any reason, continue without image blocking.
        pass
    context.set_extra_http_headers({"Accept-Language": "id-ID,id;q=0.9,en;q=0.8"})
    page = context.new_page()
    return page, browser, pw, ip_display


def _rumah123_show_click_zone_overlay(page) -> None:
    """Inject a visible overlay on the page showing the click zone and each click point (visible to human eye)."""
    x_min = RUMAH123_VERIFICATION_CLICK_X_MIN
    x_max = RUMAH123_VERIFICATION_CLICK_X_MAX
    y_min = RUMAH123_VERIFICATION_CLICK_Y_MIN
    y_max = RUMAH123_VERIFICATION_CLICK_Y_MAX
    step = RUMAH123_VERIFICATION_CLICK_STEP
    try:
        page.evaluate(
            """
            ([xMin, xMax, yMin, yMax, step]) => {
                const id = 'rumah123-click-zone-overlay';
                if (document.getElementById(id)) return;
                const wrap = document.createElement('div');
                wrap.id = id;
                wrap.style.cssText = 'position:fixed;left:0;top:0;width:100%;height:100%;pointer-events:none;z-index:2147483647;';
                const box = document.createElement('div');
                box.style.cssText = 'position:absolute;left:' + xMin + 'px;top:' + yMin + 'px;width:' + (xMax - xMin) + 'px;height:' + (yMax - yMin) + 'px;border:3px solid rgba(255,80,0,0.95);background:rgba(255,180,0,0.25);box-sizing:border-box;';
                wrap.appendChild(box);
                for (let y = yMin; y <= yMax; y += step) {
                    for (let x = xMin; x <= xMax; x += step) {
                        const dot = document.createElement('div');
                        dot.style.cssText = 'position:absolute;left:' + (x - 2) + 'px;top:' + (y - 2) + 'px;width:4px;height:4px;border-radius:2px;background:rgba(255,0,0,0.8);';
                        wrap.appendChild(dot);
                    }
                }
                document.body.appendChild(wrap);
                setTimeout(() => { const e = document.getElementById(id); if (e) e.remove(); }, 8000);
            }
            """,
            [x_min, x_max, y_min, y_max, step],
        )
    except Exception as e:
        log.debug("rumah123 | show click zone overlay failed: %s", e)


def _rumah123_click_verification_by_position(page) -> bool:
    """Simulate a wave of clicks across the verification widget area to hit the checkbox. Returns True if any click was sent. Max clicks per wave = RUMAH123_VERIFICATION_CLICK_MAX."""
    try:
        step = RUMAH123_VERIFICATION_CLICK_STEP
        x_min, x_max = RUMAH123_VERIFICATION_CLICK_X_MIN, RUMAH123_VERIFICATION_CLICK_X_MAX
        y_min, y_max = RUMAH123_VERIFICATION_CLICK_Y_MIN, RUMAH123_VERIFICATION_CLICK_Y_MAX
        max_clicks = RUMAH123_VERIFICATION_CLICK_MAX
        url_before = ""
        try:
            url_before = page.url or ""
        except Exception:
            url_before = ""
        _rumah123_show_click_zone_overlay(page)
        # Minimal pause so overlay renders before the click wave starts
        time.sleep(0.3)
        count = 0
        for y in range(y_min, y_max + 1, step):
            for x in range(x_min, x_max + 1, step):
                if count >= max_clicks:
                    break
                try:
                    page.mouse.click(x, y)
                    count += 1
                    # Slow the pace further (~50%% slower than previous 0.06s)
                    time.sleep(0.09)
                except Exception:
                    pass
            if count >= max_clicks:
                break
        if count:
            log.info("rumah123 | verification area: sent %s clicks across widget region", count)
            # Allow the challenge widget and redirect to settle after the full click wave:
            # wait until the URL changes (challenge solved / redirect), but cap at 16s.
            wait_start = time.time()
            while time.time() - wait_start < 16:
                try:
                    current_url = page.url or ""
                except Exception:
                    current_url = ""
                if url_before and current_url and current_url != url_before:
                    log.debug("rumah123 | verification page URL changed after clicks → continuing")
                    break
                time.sleep(0.5)
            return True
    except Exception as e:
        log.debug("rumah123 | verification wave click failed: %s", e)
    return False


def _rumah123_cloudflare_click_if_present(page, headless: bool = True, timeout_ms: int = 6000) -> bool:
    """If a Cloudflare challenge / Rumah123 security page is visible, try to click the checkbox. Returns True if clicked.
    Position-based clicks are only sent when we're on a verification page (content check) and in visible (non-headless) mode."""
    try:
        time.sleep(1.5)
        # 0) Wave of clicks ONLY on verification page and only in visible browser
        if not headless and _rumah123_page_has_verification(page.content().lower()):
            _rumah123_click_verification_by_position(page)
        # 1) Try frame/label-based click (if checkbox is in an iframe we can access)
        # 1) Look through all frames for the Cloudflare / Rumah123 challenge.
        for frame in page.frames:
            url = (frame.url or "").lower()
            if "challenges.cloudflare.com" not in url and "captcha-delivery" not in url and "cloudflare" not in url:
                # Still consider the frame if it contains the Indonesian verification text.
                try:
                    if not frame.is_detached():
                        if not frame.locator("body").is_visible(timeout=500):
                            continue
                except Exception:
                    continue
            try:
                # Check for the Indonesian label text inside the frame.
                label = frame.get_by_text("Buktikan bahwa Anda adalah manusia", exact=False).first
                if label.is_visible(timeout=800):
                    # Try Playwright's label/checkbox helpers first.
                    try:
                        checkbox = frame.get_byLabel("Buktikan bahwa Anda adalah manusia", exact=False).first  # type: ignore[attr-defined]
                        if checkbox.is_visible(timeout=800):
                            checkbox.click()
                            log.info("rumah123 | clicked Rumah123 verification checkbox via label in frame %s", url or "<no-url>")
                            time.sleep(3)
                            return True
                    except Exception:
                        pass
                    # Fallback: try common checkbox-like selectors near the label.
                    for cb_sel in (
                        "input[type='checkbox']",
                        "[role='checkbox']",
                        ".mark",
                        ".cf-turnstile",
                        "div[aria-checked]",
                        "label:has-text('Buktikan bahwa Anda adalah manusia')",
                    ):
                        try:
                            el = frame.locator(cb_sel).first
                            if el.is_visible(timeout=800):
                                el.click()
                                log.info("rumah123 | clicked Rumah123 verification checkbox via selector '%s' in frame %s", cb_sel, url or "<no-url>")
                                time.sleep(3)
                                return True
                        except Exception:
                            continue
            except Exception:
                continue

        # 2) Fallback: click by visible text on main page (if no iframe/frame match)
        for text in ("Buktikan bahwa Anda adalah manusia", "Verify you are human", "I am human", "I'm not a robot"):
            try:
                btn = page.get_by_text(text, exact=False).first
                if btn.is_visible(timeout=1000):
                    btn.click()
                    log.info("rumah123 | clicked Cloudflare challenge (text)")
                    time.sleep(4)
                    return True
            except Exception:
                continue
    except Exception as e:
        log.debug("rumah123 | cloudflare auto-click failed: %s", e)
    return False


def _rumah123_fetch_listing_in_visible_browser(page_num: int, close_before: tuple | None = None) -> tuple[list[str], bool]:
    """Open a temporary non-headless browser, load the listing page, apply clicks if verification appears, collect links, then close browser.
    If close_before is (browser, pw), close them first so only one Chromium runs. Returns (links, ok)."""
    if close_before is not None:
        try:
            browser, pw = close_before
            browser.close()
            pw.stop()
            log.info("rumah123 | closed headless browser before opening visible one")
        except Exception as e:
            log.debug("rumah123 | closing headless before visible: %s", e)
    url = RUMAH123_LISTING if page_num <= 1 else f"{RUMAH123_LISTING}?page={page_num}"
    for attempt in range(2):
        log.warning("rumah123 | verification required: opening temporary VISIBLE browser for listing page %s (attempt %s)", page_num, attempt + 1)
        handle = _rumah123_browser(headless=False)
        if not handle:
            log.error("rumah123 | failed to launch visible browser (playwright/chromium may be missing); cannot complete verification")
            return ([], False)
        page, browser, pw, _ = handle
        log.info("rumah123 | visible browser launched; loading page and applying clicks in widget area")
        session_start = time.time()
        try:
            page.set_extra_http_headers(_rumah123_headers())
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            # Small settle time, then immediately begin verification checks and click wave if needed.
            time.sleep(2)
            start = time.time()
            # Verification loop; also respect a hard 30s cap for this Chromium session
            while _rumah123_page_has_verification(page.content().lower()) and (time.time() - start) < 10:
                if time.time() - session_start > 30:
                    log.warning("rumah123 | listing visible session >30s on page %s; aborting and relaunching", page_num)
                    break
                _rumah123_cloudflare_click_if_present(page, headless=False)
                time.sleep(3)
            # After clicks, if the challenge text is still present, kill and relaunch once.
            html = page.content()
            html_lower = html.lower()
            if "verifikasi bahwa anda adalah manusia. ini dapat memerlukan waktu beberapa detik." in html_lower or (time.time() - session_start > 30):
                log.warning("rumah123 | listing challenge text still present after clicks; closing and relaunching (attempt %s)", attempt + 1)
                continue
            try:
                page.wait_for_selector("a[href*='/properti/']", timeout=8_000)
            except Exception:
                pass
            time.sleep(2)
            html = page.content()
            soup = BeautifulSoup(html, "html.parser")
            links: list[str] = []
            for a in soup.find_all("a", href=True):
                href = (a.get("href") or "").strip()
                if not href.startswith("/properti/") or "/agent-" in href or "/independent-property-agent/" in href:
                    continue
                if re.search(r"-(?:hor|vlr)\d+/?$", href):
                    path = href.rstrip("/") + "/" if not href.endswith("/") else href
                    full = urljoin(RUMAH123_BASE, path)
                    if full not in links:
                        links.append(full)
            elapsed = time.time() - session_start
            log.info("rumah123 | visible browser closed; resuming headless (collected %s links from page %s, elapsed %.1fs)", len(links), page_num, elapsed)
            return (links, bool(links))
        finally:
            browser.close()
            pw.stop()
    log.warning("rumah123 | visible listing challenge could not be cleared after retries; giving up on page %s", page_num)
    return ([], False)


def _rumah123_fetch_detail_in_visible_browser(url: str, close_before: tuple | None = None) -> tuple[dict | None, bool]:
    """Open a temporary non-headless browser, load the detail URL, apply clicks if verification appears, parse row, then close browser.
    If close_before is (browser, pw), close them first so only one Chromium runs (avoids visible window not opening on macOS). Returns (row, ok)."""
    if close_before is not None:
        try:
            browser, pw = close_before
            browser.close()
            pw.stop()
            log.info("rumah123 | closed headless browser before opening visible one")
        except Exception as e:
            log.debug("rumah123 | closing headless before visible: %s", e)
    for attempt in range(2):
        log.warning("rumah123 | verification required: opening temporary VISIBLE browser for detail page (attempt %s)", attempt + 1)
        handle = _rumah123_browser(headless=False)
        if not handle:
            log.error("rumah123 | failed to launch visible browser (playwright/chromium may be missing); cannot complete verification")
            return (None, False)
        page, browser, pw, _ = handle
        log.info("rumah123 | visible browser launched; loading page and applying clicks in widget area")
        session_start = time.time()
        try:
            page.set_extra_http_headers(_rumah123_headers())
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            # Small settle time, then immediately begin verification checks and click wave if needed.
            time.sleep(2)
            start = time.time()
            # Verification loop; also respect a hard 30s cap for this Chromium session
            while _rumah123_page_has_verification(page.content().lower()) and (time.time() - start) < 10:
                if time.time() - session_start > 30:
                    log.warning("rumah123 | detail visible session >30s for %s; aborting and relaunching", url)
                    break
                _rumah123_cloudflare_click_if_present(page, headless=False)
                time.sleep(3)
            html = page.content()
            html_lower = html.lower()
            if "verifikasi bahwa anda adalah manusia. ini dapat memerlukan waktu beberapa detik." in html_lower or (time.time() - session_start > 30):
                log.warning("rumah123 | detail challenge text still present after clicks; closing and relaunching (attempt %s)", attempt + 1)
                continue
            try:
                page.wait_for_selector("h1", timeout=8_000)
            except Exception:
                pass
            time.sleep(1)
            html = page.content()
            soup = BeautifulSoup(html, "html.parser")
            row = _rumah123_soup_to_row(soup, url, html=html)
            elapsed = time.time() - session_start
            log.info("rumah123 | visible browser closed; resuming headless (elapsed %.1fs for %s)", elapsed, url)
            return (row, row is not None)
        finally:
            browser.close()
            pw.stop()
    log.warning("rumah123 | visible detail challenge could not be cleared after retries; giving up on %s", url)
    return (None, False)


def _rumah123_fetch_listing_playwright(
    page, page_num: int, listing_sleep: float, headless: bool = True, browser_handle: tuple | None = None
) -> tuple[list[str], bool, bool, bool]:
    """Fetch one listing page via Playwright. Returns (links, ok, security_blocked, browser_closed)."""
    url = RUMAH123_LISTING if page_num <= 1 else f"{RUMAH123_LISTING}?page={page_num}"
    security_blocked = False
    close_before = (browser_handle[1], browser_handle[2]) if browser_handle else None
    security_wait_sec = RUMAH123_SECURITY_RETRY_SEC if headless else 120
    for attempt in range(3):
        try:
            page.set_extra_http_headers(_rumah123_headers())
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            time.sleep(1)  # shorter wait: let verification challenge/render settle
            html_lower = page.content().lower()
            if headless and _rumah123_page_has_verification(html_lower):
                log.info("rumah123 | verification detected on listing page %s (initial check); opening visible browser", page_num)
                links, ok = _rumah123_fetch_listing_in_visible_browser(page_num, close_before=close_before)
                return (links, ok, False, close_before is not None)
            start = time.time()
            while True:
                _rumah123_cloudflare_click_if_present(page, headless=headless)
                html_lower = page.content().lower()
                if not _rumah123_page_has_verification(html_lower):
                    break
                if headless:
                    links, ok = _rumah123_fetch_listing_in_visible_browser(page_num, close_before=close_before)
                    return (links, ok, False, close_before is not None)
                if time.time() - start > security_wait_sec:
                    log.warning("rumah123 | verification still present after %ss on listing page %s; click the checkbox in the browser window", security_wait_sec, page_num)
                    break
                log.info("rumah123 | waiting for security verification (click the checkbox in the browser if visible)...")
                time.sleep(5)

            if page_num == 1:
                log.info("rumah123 | waiting for property links after any security verification...")
            try:
                # Slightly tighter timeout so slow pages don't hold up the whole crawl
                page.wait_for_selector("a[href*='/properti/']", timeout=8_000)
            except Exception:
                if headless:
                    log.info("rumah123 | listing page %s: no property links found in time; assuming verification, opening visible browser", page_num)
                    links, ok = _rumah123_fetch_listing_in_visible_browser(page_num, close_before=close_before)
                    return (links, ok, False, close_before is not None)
                raise
            time.sleep(2)
            html = page.content()
            break
        except Exception as e:
            if _is_proxy_error(e):
                _switch_to_proxyless()
                return ([], False, False, True)
            if attempt < 2:
                time.sleep(5 * (attempt + 1))
            else:
                log.warning("rumah123 | page %s failed after retries: %s", page_num, e)
                return ([], False, False, False)
    else:
        return ([], False, False, False)
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href.startswith("/properti/") or "/agent-" in href or "/independent-property-agent/" in href:
            continue
        if re.search(r"-(?:hor|vlr)\d+/?$", href):
            path = href.rstrip("/") + "/" if not href.endswith("/") else href
            full = urljoin(RUMAH123_BASE, path)
            if full not in links:
                links.append(full)
    return (links, True, False, False)


def _rumah123_fetch_detail_playwright(
    page, url: str, delay: float, headless: bool = True, browser_handle: tuple | None = None
) -> tuple[dict | None, bool, bool]:
    """Fetch one detail page via Playwright. Returns (parsed_row_or_none, security_blocked, browser_closed).
    When browser_closed is True, caller must re-acquire headless browser (we closed it to open visible).
    Overall per-URL safety timeout is ~30s: after that we give up and let the caller continue with next URL."""
    time.sleep(delay)
    overall_start = time.time()
    security_wait_sec = RUMAH123_SECURITY_RETRY_SEC if headless else 120
    close_before = (browser_handle[1], browser_handle[2]) if browser_handle else None  # (browser, pw)
    for attempt in range(3):
        # Global safety break: if this URL has already taken ~30s, abort and resume on next one.
        if time.time() - overall_start > 30:
            log.warning("rumah123 | detail fetch timeout after 30s for %s; skipping and resuming", url)
            return (None, False, False)
        try:
            page.set_extra_http_headers(_rumah123_headers())
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            time.sleep(3)  # let verification challenge render (can appear after a delay)
            html_lower = page.content().lower()
            # In headless, as soon as we see the challenge: close headless, open visible, collect row, then caller re-acquires headless
            if headless and _rumah123_page_has_verification(html_lower):
                log.info("rumah123 | verification detected on detail page (initial check); opening visible browser")
                row, _ = _rumah123_fetch_detail_in_visible_browser(url, close_before=close_before)
                return (row, False, close_before is not None)
            start = time.time()
            while True:
                _rumah123_cloudflare_click_if_present(page, headless=headless)
                html_lower = page.content().lower()
                if not _rumah123_page_has_verification(html_lower):
                    break
                if headless:
                    row, _ = _rumah123_fetch_detail_in_visible_browser(url, close_before=close_before)
                    return (row, False, close_before is not None)
                if time.time() - start > security_wait_sec:
                    log.warning("rumah123 | verification still present after %ss; click the checkbox in the browser window", security_wait_sec)
                    break
                log.info("rumah123 | waiting for security verification (click the checkbox in the browser if visible)...")
                time.sleep(5)
            try:
                page.wait_for_selector("h1", timeout=15_000)
            except Exception:
                if headless:
                    log.info("rumah123 | detail page: no h1 found in time; assuming verification, opening visible browser")
                    row, _ = _rumah123_fetch_detail_in_visible_browser(url, close_before=close_before)
                    return (row, False, close_before is not None)
                raise
            time.sleep(1)
            html = page.content()
            soup = BeautifulSoup(html, "html.parser")
            return (_rumah123_soup_to_row(soup, url, html=html), False, False)
        except Exception as e:
            if _is_proxy_error(e):
                _switch_to_proxyless()
                return (None, False, True)
            if attempt < 2:
                time.sleep(4 * (attempt + 1))
            else:
                log.debug("rumah123 | error %s: %s", url[:60], e)
                return (None, False, False)
    return (None, False, False)


# ---------------------------------------------------------------------------
# Villa-Bali
# ---------------------------------------------------------------------------

VILLA_BALI_BASE = "https://www.villa-bali.com"
VILLA_BALI_SEARCH = f"{VILLA_BALI_BASE}/en/search"
USD_TO_IDR = 16_000


def _human_delay(base: float, jitter: float = 2.0):
    time.sleep(base + random.uniform(0, jitter))


def _usd_to_idr(usd: float) -> str:
    return f"Rp {int(round(usd * USD_TO_IDR)):,}".replace(",", ".")


def villa_bali_parse_price(price_text: str) -> tuple[str, str]:
    m = re.search(r"USD\s*[\d,]+(?:\.\d+)?", (price_text or ""), re.I)
    if not m:
        return "", "day"
    num_str = re.sub(r"[^\d.]", "", m.group(0))
    try:
        return _usd_to_idr(float(num_str)), "day"
    except ValueError:
        return "", "day"


def _villa_bali_headers() -> dict:
    return {
        "User-Agent": _random_user_agent(),
        "Accept-Language": "en-US,en;q=0.9",
    }


def villa_bali_get_browser(headless: bool = False):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise SystemExit("Install playwright: pip install playwright && playwright install chromium")
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=headless, args=["--disable-blink-features=AutomationControlled", "--no-sandbox"])
    proxy, ip_display = _next_proxy_for_playwright()
    log.info("villa-bali | browser using %s", ip_display)
    context_options = {
        "user_agent": _random_user_agent(),
        "locale": "en-US",
        "viewport": {"width": 1920, "height": 1080},
    }
    if proxy is not None:
        context_options["proxy"] = proxy
    context = browser.new_context(**context_options)
    # Do not load image contents to reduce bandwidth and speed up scraping.
    try:
        context.route(
            "**/*",
            lambda route, request: route.abort()
            if request.resource_type == "image"
            else route.continue_(),
        )
    except Exception:
        pass
    context.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
    page = context.new_page()
    return page, browser, pw


def villa_bali_fetch_listing_links(page) -> list[str]:
    _human_delay(2, 2)
    page.set_extra_http_headers(_villa_bali_headers())
    page.goto(VILLA_BALI_SEARCH, wait_until="load", timeout=45000)
    _human_delay(5, 4)
    for _ in range(5):
        page.evaluate("window.scrollBy(0, 800)")
        _human_delay(1.2, 1.5)
    page.evaluate("window.scrollTo(0, 0)")
    _human_delay(1.5, 1)
    soup = BeautifulSoup(page.content(), "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if "/en/villa/" in href:
            full = urljoin(VILLA_BALI_BASE, href)
            if full not in links:
                links.append(full)
    return links


def villa_bali_is_blocked(html: str) -> bool:
    lower = html.lower()
    return "you have been blocked" in lower or "captcha" in lower or "access denied" in lower


def villa_bali_parse_detail(page, url: str, delay: float = 1.0) -> dict | None:
    _human_delay(delay, 4)
    for attempt in range(2):
        try:
            page.set_extra_http_headers(_villa_bali_headers())
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
        except Exception:
            if attempt == 0:
                _human_delay(15, 10)
            continue
        _human_delay(2, 2)
        html = page.content()
        if villa_bali_is_blocked(html):
            if attempt == 0:
                _human_delay(20, 15)
            continue
        break
    else:
        return None
    soup = BeautifulSoup(html, "html.parser")
    h1 = soup.find("h1")
    title = h1.get_text(strip=True).replace(",", "") if h1 else ""
    price_idr = ""
    duration = "day"
    for el in soup.find_all(string=re.compile(r"USD\s*\d|per night|per week", re.I)):
        t = el.strip() if isinstance(el, str) else el
        if "USD" in t:
            price_idr, duration = villa_bali_parse_price(t)
            break
    location = ""
    for el in soup.find_all(string=re.compile(r"Candidasa|Jimbaran|Seminyak|Canggu|Ubud|Sanur|Pererenan|Seseh|Amed|Nusa Dua|The Bukit", re.I)):
        t = el.strip() if isinstance(el, str) else el
        if 2 <= len(t) <= 80:
            location = t.replace(",", "")
            break
    bedrooms = ""
    bathrooms = ""
    for el in soup.find_all(string=re.compile(r"\d+\s*bedroom|\d+\s*bathroom", re.I)):
        t = el.strip() if isinstance(el, str) else el
        m = re.search(r"(\d+)\s*bedroom", t, re.I)
        if m:
            bedrooms = m.group(1)
        m = re.search(r"(\d+)\s*bathroom", t, re.I)
        if m:
            bathrooms = m.group(1)
    description = ""
    desc_el = soup.find(class_=re.compile(r"description|overview|about", re.I))
    if desc_el:
        description = desc_el.get_text(separator=" ", strip=True).replace(",", "")[:5000]
    if not description:
        for p in soup.find_all("p"):
            t = p.get_text(strip=True)
            if len(t) > 100:
                description = t.replace(",", "")[:5000]
                break
    facilities = ""
    fac_el = soup.find(class_=re.compile(r"facilit|amenit|feature", re.I))
    if fac_el:
        parts = [n.get_text(strip=True).replace(",", "") for n in fac_el.find_all(["li", "span", "div"]) if 2 < len(n.get_text(strip=True)) < 60]
        facilities = " | ".join(parts[:30]) if parts else ""
    main_image = ""
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        main_image = (og["content"] or "").strip()
    return {
        "url": url,
        "main_image": main_image,
        "title": title,
        "price_idr": price_idr,
        "duration": duration,
        "location": location,
        "bedrooms": bedrooms,
        "bathrooms": bathrooms,
        "land_size_m2": "",
        "building_size_m2": "",
        "certificate": "",
        "description": description,
        "facilities": facilities,
        "agent_name": "Villa-Bali.com",
        "updated_by": "",
    }


# ---------------------------------------------------------------------------
# Resume & duplicate detection (load existing URLs from CSV/Excel)
# ---------------------------------------------------------------------------

# Query params we strip so the same listing is one key (e.g. ?ref=... or utm_*)
_TRACKING_QUERY_PARAMS = frozenset(
    {"ref", "utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term"}
)


def _normalize_listing_url(url: str) -> str:
    """
    Normalize a listing URL for resume/duplicate detection: strip fragment and
    tracking query params so the same listing is never treated as two.
    """
    u = (url or "").strip()
    if not u:
        return ""
    try:
        parsed = urlparse(u)
        # Drop fragment
        netloc = parsed.netloc or ""
        path = parsed.path.rstrip("/") or "/"
        # Keep query but drop tracking params
        q = parse_qs(parsed.query, keep_blank_values=False)
        for key in list(q):
            if key.lower() in _TRACKING_QUERY_PARAMS:
                del q[key]
        query = urlencode(q, doseq=True) if q else ""
        return urlunparse((parsed.scheme, netloc, path, parsed.params, query, ""))
    except Exception:
        return u


def _load_existing_urls(path: Path) -> set[str]:
    """
    Load all existing listing URLs from the output file (CSV or Excel) for resume and duplicate detection.
    Returns a set of *normalized* URL strings. Same listing with different ?ref= or fragment = one entry.
    Returns empty set if file missing or unreadable.
    """
    if not path.exists():
        return set()
    urls = set()
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            with open(path, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames and "url" in reader.fieldnames:
                    for row in reader:
                        u = (row.get("url") or "").strip()
                        if u:
                            urls.add(_normalize_listing_url(u))
        elif suffix in (".xlsx", ".xls"):
            try:
                import openpyxl
            except ImportError:
                log.debug("openpyxl not installed; cannot read Excel for existing URLs")
                return set()
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb.active
            if sheet is None:
                wb.close()
                return set()
            header = None
            url_col = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header = row
                break
            if header:
                for i, h in enumerate(header):
                    if (h or "").strip().lower() == "url":
                        url_col = i
                        break
            if url_col is not None:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > url_col:
                        u = (row[url_col] or "").strip()
                        if isinstance(u, str) and u:
                            urls.add(_normalize_listing_url(u))
            wb.close()
    except Exception as e:
        log.warning("Could not load existing URLs from %s: %s", path, e)
    return urls


# Lock for thread-safe CSV writes and existing_urls updates when using parallel detail workers.
_csv_write_lock = threading.Lock()


def _write_row_to_csv(
    csv_handle: tuple | None,
    row: dict,
    existing_urls: set[str] | None = None,
) -> bool:
    """
    If csv_handle is (writer, file), write one row and flush.
    If existing_urls is set and row["url"] (normalized) is already in it, skip write and return False.
    Otherwise write, add normalized url to existing_urls, and return True. Prevents double entries.
    Thread-safe when multiple workers call this concurrently.
    """
    if csv_handle is None:
        return True
    with _csv_write_lock:
        url = (row.get("url") or "").strip()
        key = _normalize_listing_url(url) if url else ""
        if existing_urls is not None and key and key in existing_urls:
            return False
        writer, f = csv_handle
        writer.writerow(row)
        f.flush()
        if existing_urls is not None and key:
            existing_urls.add(key)
        return True


def run_rumah123(
    args,
    csv_handle: tuple | None = None,
    existing_urls: set[str] | None = None,
) -> list[dict]:
    log.info("rumah123 | fetching listing pages...")
    seen = existing_urls if existing_urls is not None else set()
    listing_sleep = 0.5
    base_delay = getattr(args, "delay", 3.0)
    if getattr(args, "fast", False):
        base_delay = max(1.0, base_delay * 0.55)

    # Stage 1: always start headless; when a page shows verification it is re-opened in a one-off visible browser, then we resume headless
    headless = getattr(args, "headless", True)
    browser_handle = _rumah123_browser(headless=headless)
    if browser_handle is not None:
        page, browser, pw, current_proxy = browser_handle
        try:
            log.info("rumah123 | launching browser (headless=%s)", headless)
            listing_batch_size = 10
            max_listing_pages = max(1, args.pages)
            # Heuristic resume: estimate how many listing pages were already covered from the CSV.
            # Assume ~20 listings per page; resume from page ~= (existing_rows / 20) + 1.
            AVG_LISTINGS_PER_PAGE = 20
            existing_rows = len(seen)
            if existing_rows:
                approx_pages_done = max(0, existing_rows // AVG_LISTINGS_PER_PAGE)
                resume_page = min(max_listing_pages, approx_pages_done + 1)
            else:
                resume_page = 1
            if resume_page > 1:
                log.info(
                    "rumah123 | resume: %s existing rows → starting listing pagination from page %s (max %s)",
                    existing_rows,
                    resume_page,
                    max_listing_pages,
                )
            detected_total_pages: int | None = None
            p = resume_page
            all_rows: list[dict] = []
            if args.list_only:
                all_links_list: list[str] = []
            while p <= max_listing_pages:
                batch_end = min(p + listing_batch_size - 1, max_listing_pages)
                log.info("rumah123 | batch: listing pages %s–%s (of max %s)", p, batch_end, max_listing_pages)
                batch_links: list[str] = []
                page_num = p
                while page_num <= batch_end:
                    if detected_total_pages:
                        log.info("rumah123 | requesting listing page %s/%s via proxy %s", page_num, detected_total_pages, current_proxy)
                    else:
                        log.info("rumah123 | requesting listing page %s (total unknown yet) via proxy %s", page_num, current_proxy)
                    if page_num > 1:
                        time.sleep(listing_sleep)
                    links, ok, _, browser_closed = _rumah123_fetch_listing_playwright(
                        page, page_num, listing_sleep, headless=headless, browser_handle=(page, browser, pw)
                    )
                    if browser_closed:
                        browser_handle = _rumah123_browser(headless=headless)
                        if not browser_handle:
                            log.error("rumah123 | failed to re-launch headless browser after visible fetch")
                            break
                        page, browser, pw, current_proxy = browser_handle
                        log.info("rumah123 | re-launched headless browser (proxy %s)", current_proxy)
                    if not ok:
                        time.sleep(8)
                        page_num += 1
                        continue
                    if page_num == 1 and detected_total_pages is None:
                        try:
                            html = page.content()
                            detected = _rumah123_extract_max_pages_from_html(html)
                        except Exception:
                            detected = None
                        if detected and detected > 0:
                            detected_total_pages = detected
                            max_listing_pages = min(args.pages, detected_total_pages)
                            log.info(
                                "rumah123 | pagination: detected %s total listing pages; will fetch up to %s (CLI limit %s)",
                                detected_total_pages,
                                max_listing_pages,
                                args.pages,
                            )
                    if not links:
                        break
                    for link in links:
                        if link not in batch_links:
                            batch_links.append(link)
                    page_num += 1
                log.info("rumah123 | batch: collected %s unique links from pages %s–%s", len(batch_links), p, batch_end)
                if args.list_only:
                    all_links_list.extend(batch_links)
                    p = batch_end + 1
                    continue
                links_to_fetch = [u for u in batch_links if _normalize_listing_url(u) not in seen]
                skipped = len(batch_links) - len(links_to_fetch)
                if skipped:
                    log.info("rumah123 | batch: skipping %s URLs already in output", skipped)
                if links_to_fetch:
                    delay = base_delay
                    total_detail = len(links_to_fetch)
                    log.info(
                        "rumah123 | batch: fetching %s detail pages in sub-batches of 6 (delay ~%ss)",
                        total_detail,
                        round(delay, 1),
                    )
                    # Process detail pages in groups of 6 (still sequential, but logged 6-by-6)
                    done = 0
                    for i in range(0, total_detail, 6):
                        sub_batch = links_to_fetch[i : i + 6]
                        log.debug("rumah123 | detail sub-batch %s-%s/%s", i + 1, i + len(sub_batch), total_detail)
                        for url in sub_batch:
                            done += 1
                            if done % 50 == 0 or done == total_detail:
                                log.info("rumah123 | detail progress %s/%s", done, total_detail)
                            log.info("rumah123 | detail page %s via proxy %s", url, current_proxy)
                            row, _, browser_closed = _rumah123_fetch_detail_playwright(
                                page, url, delay, headless=headless, browser_handle=(page, browser, pw)
                            )
                            if browser_closed:
                                browser_handle = _rumah123_browser(headless=headless)
                                if not browser_handle:
                                    log.error("rumah123 | failed to re-launch headless browser after visible fetch")
                                    break
                                page, browser, pw, current_proxy = browser_handle
                                log.info("rumah123 | re-launched headless browser (proxy %s)", current_proxy)
                            if row and _row_has_title_and_price(row, "rumah123"):
                                flat = flatten_row(row, source="rumah123")
                                if _write_row_to_csv(csv_handle, flat, seen):
                                    all_rows.append(flat)
                p = batch_end + 1
            if args.list_only:
                return list(dict.fromkeys(all_links_list))
            return all_rows
        finally:
            try:
                browser.close()
                pw.stop()
            except Exception:
                pass

    # Fallback: requests (no browser, e.g. Playwright not installed)
    session = _rumah123_session()
    all_links = []
    pages_fetched = 0
    # Same heuristic resume for the requests-based fallback.
    AVG_LISTINGS_PER_PAGE = 20
    existing_rows = len(seen)
    if existing_rows:
        approx_pages_done = max(0, existing_rows // AVG_LISTINGS_PER_PAGE)
        start_page = min(args.pages, approx_pages_done + 1)
    else:
        start_page = 1
    if start_page > 1:
        log.info(
            "rumah123 | resume (requests fallback): %s existing rows → starting from listing page %s (max %s)",
            existing_rows,
            start_page,
            args.pages,
        )
    for p in range(start_page, args.pages + 1):
        log.info("rumah123 | requesting listing page %s", p)
        if p > 1:
            time.sleep(listing_sleep)
        try:
            links, ok = rumah123_fetch_listing_links(session, page=p)
        except Exception as e:
            log.warning("rumah123 | page %s failed: %s", p, e)
            continue
        if not ok:
            time.sleep(8)
            continue
        if not links:
            break
        before = len(all_links)
        for link in links:
            if link not in all_links:
                all_links.append(link)
        pages_fetched += 1
        if len(all_links) == before:
            break
    log.info("rumah123 | collected %s unique property URLs from %s page(s)", len(all_links), pages_fetched)
    if not all_links:
        log.warning("rumah123 | no listing links found")
        return []
    if args.list_only:
        return all_links
    links_to_fetch = [u for u in all_links if u not in seen]
    skipped = len(all_links) - len(links_to_fetch)
    if skipped:
        log.info("rumah123 | skipping %s URLs already in output (resume/duplicate detection)", skipped)
    if not links_to_fetch:
        log.info("rumah123 | no new URLs to fetch")
        return []
    delay = base_delay
    workers = max(1, getattr(args, "workers", 3))
    workers = min(workers, len(links_to_fetch))
    rows = []

    def _rumah123_fetch_one(url: str) -> tuple[str, dict | None]:
        try:
            s = _rumah123_session()
            return (url, rumah123_parse_detail(s, url, delay=delay))
        except Exception as e:
            log.warning("rumah123 | error %s: %s", url, e)
            return (url, None)

    if workers <= 1:
        log.info("rumah123 | fetching %s detail pages sequentially (delay ~%ss)", len(links_to_fetch), round(delay, 1))
        total = len(links_to_fetch)
        for done, url in enumerate(links_to_fetch, 1):
            if done % 50 == 0 or done == total:
                log.info("rumah123 | detail progress %s/%s", done, total)
            _, row = _rumah123_fetch_one(url)
            if row and _row_has_title_and_price(row, "rumah123"):
                log.debug("rumah123 | ok %s", url)
                flat = flatten_row(row, source="rumah123")
                if _write_row_to_csv(csv_handle, flat, seen):
                    rows.append(flat)
            elif row:
                log.debug("rumah123 | skip (no title/price) %s", url)
    else:
        log.info("rumah123 | fetching %s detail pages with %s workers (delay ~%ss)", len(links_to_fetch), workers, round(delay, 1))
        total = len(links_to_fetch)
        done_count = [0]

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_url = {executor.submit(_rumah123_fetch_one, u): u for u in links_to_fetch}
            for future in as_completed(future_to_url):
                url, row = future.result()
                with _csv_write_lock:
                    done_count[0] += 1
                    n = done_count[0]
                if n % 50 == 0 or n == total:
                    log.info("rumah123 | detail progress %s/%s", n, total)
                if row and _row_has_title_and_price(row, "rumah123"):
                    flat = flatten_row(row, source="rumah123")
                    if _write_row_to_csv(csv_handle, flat, seen):
                        rows.append(flat)
                elif row:
                    log.debug("rumah123 | skip (no title/price) %s", url)
    return rows


###############################################################################
# OLX Badung (rent houses/apartments)
###############################################################################

OLX_BADUNG_BASE = "https://www.olx.co.id/kab-badung_g4000217/q-villa-rental"


def _olx_headers() -> dict:
    # Reuse random desktop UA pool
    return {
        "User-Agent": _random_user_agent(),
        "Accept-Language": "id-ID,id;q=0.9,en;q=0.8",
    }

def olx_get_browser(login_wait: bool = True):
    """
    Launch a visible Chromium browser for OLX.

    Headless is explicitly disabled here because OLX quickly blocks headless
    automation. Interact with the visible window if a verification step appears.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        raise SystemExit(
            "Install playwright: pip install playwright && playwright install chromium chrome"
        )
    pw = sync_playwright().start()

    # Prefer the locally installed Google Chrome (visible window), falling
    # back to the bundled Chromium if the Chrome channel is not available.
    try:
        browser = pw.chromium.launch(
            channel="chrome",
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
    except Exception:
        browser = pw.chromium.launch(
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
    proxy, ip_display = _next_proxy_for_playwright()
    log.info("olx-badung | browser using %s", ip_display)
    context_options = {
        "user_agent": _random_user_agent(),
        "locale": "id-ID",
        "viewport": {"width": 1920, "height": 1080},
    }
    if proxy is not None:
        context_options["proxy"] = proxy
    context = browser.new_context(**context_options)
    # Let CSS, images, and fonts load so the list page renders fully and the
    # "Muat lainnya" button and layout are visible.
    context.set_extra_http_headers({"Accept-Language": "id-ID,id;q=0.9,en;q=0.8"})
    page = context.new_page()

    # Immediately open OLX in this page so any manual login/captcha you do
    # is stored in the SAME browser context and will be reused for all
    # subsequent listing/detail requests.
    try:
        log.info("olx-badung | opening OLX home in Chrome for manual login (session will be reused for scraping)")
        page.goto("https://www.olx.co.id/", wait_until="domcontentloaded", timeout=45000)
    except Exception as e:
        log.warning("olx-badung | initial OLX home navigation failed: %s", e)

    # Give you time to complete any manual verification (e.g. login, captcha)
    # in the visible Chrome window before we start scraping, unless explicitly
    # disabled via CLI.
    if login_wait:
        log.info(
            "olx-badung | Chrome ready on OLX; waiting 60s so you can complete login/verification"
        )
        try:
            time.sleep(60)
        except KeyboardInterrupt:
            log.warning("olx-badung | 60s login wait interrupted; continuing immediately")

    return page, browser, pw


def _olx_extract_item_links_from_html(html: str) -> list[str]:
    """Parse HTML and return normalized OLX item URLs (/item/...)."""
    links: list[str] = []
    seen: set[str] = set()
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href or "/item/" not in href:
            continue
        full = urljoin("https://www.olx.co.id", href)
        try:
            parsed = urlparse(full)
            full = parsed._replace(query="").geturl()
        except Exception:
            pass
        if full not in seen:
            seen.add(full)
            links.append(full)
    return links


def _olx_wait_page_fully_loaded(page) -> None:
    """
    Wait until the OLX listing page is fully loaded before we look for the
    'Muat lainnya' button: network idle, listing content visible, and a settle
    period so the full layout (including bottom button) is rendered.
    """
    # 1) Wait for load event
    try:
        page.wait_for_load_state("load", timeout=30_000)
    except Exception as e:
        log.debug("olx-badung | wait load: %s", e)

    # 2) Wait for listing API/content: at least one item link in the DOM
    try:
        page.wait_for_selector('a[href*="/item/"]', state="visible", timeout=25_000)
    except Exception as e:
        log.debug("olx-badung | wait for first listing link: %s", e)

    # 3) Wait for network to go idle (no requests for 500ms)
    try:
        page.wait_for_load_state("networkidle", timeout=25_000)
    except Exception:
        pass

    # 4) Poll until we see a reasonable number of listing links (first batch rendered)
    min_listing_links = 5
    for _ in range(60):  # up to 30 seconds
        try:
            count = page.locator('a[href*="/item/"]').count()
            if count >= min_listing_links:
                log.info("olx-badung | listing content ready (%s item links visible)", count)
                break
        except Exception:
            pass
        time.sleep(0.5)
    else:
        log.debug("olx-badung | continuing after 30s (some listing links may be present)")

    # 5) Short settle; then caller will wait explicitly for the button to appear
    time.sleep(3)


def _olx_scroll_to_bottom(page) -> None:
    """Scroll to bottom in steps so lazy content at the bottom can render."""
    for _ in range(14):
        try:
            page.evaluate("window.scrollBy(0, 500)")
            time.sleep(0.4)
        except Exception:
            pass
    try:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1.5)
    except Exception:
        pass


OLX_LOAD_MORE_SELECTORS = (
    'button:has-text("Muat lainnya")',
    'a:has-text("Muat lainnya")',
    '[data-cy="load-more"]',
    'button:has-text("Load more")',
    'a:has-text("Load more")',
)


def _olx_is_load_more_visible(page) -> bool:
    """Return True if the load-more button is currently visible."""
    for selector in OLX_LOAD_MORE_SELECTORS:
        try:
            if page.locator(selector).first.is_visible(timeout=2000):
                return True
        except Exception:
            continue
    return False


def _olx_wait_for_load_more_button_visible(page, timeout_sec: int = 120) -> bool:
    """
    Poll: scroll to bottom, check if 'Muat lainnya' is visible. Repeat until
    the button appears or timeout_sec is reached. Return True if button was
    seen, False if timeout. Only then should we try to click it.
    """
    log.info("olx-badung | waiting for 'Muat lainnya' button to appear (polling up to %ss)...", timeout_sec)
    deadline = time.time() + timeout_sec
    attempt = 0
    while time.time() < deadline:
        attempt += 1
        _olx_scroll_to_bottom(page)
        if _olx_is_load_more_visible(page):
            log.info("olx-badung | 'Muat lainnya' button is now visible (after %s attempts)", attempt)
            return True
        if attempt % 5 == 0:
            log.info("olx-badung | still waiting for 'Muat lainnya'... attempt %s", attempt)
        time.sleep(4)
    log.warning("olx-badung | 'Muat lainnya' button did not appear after %ss", timeout_sec)
    return False


def olx_fetch_listing_links_playwright(page, max_pages: int = 1) -> list[str]:
    """
    Collect listing detail URLs from the OLX Badung category listing.

    OLX uses a single list page with lazy load: a "Muat lainnya" (Load more)
    button loads more items. We wait for the page to fully load, scroll in steps
    to reveal the button, then repeatedly click it and collect links.
    """
    links: list[str] = []
    seen: set[str] = set()

    try:
        log.info("olx-badung | loading listing page (single URL, load-more strategy)")
        page.set_extra_http_headers(_olx_headers())
        page.goto(OLX_BADUNG_BASE, wait_until="load", timeout=60_000)
        _olx_wait_page_fully_loaded(page)
        # Only proceed to click after the button has actually appeared (poll up to 2 min)
        _olx_wait_for_load_more_button_visible(page, timeout_sec=120)
    except Exception as e:
        log.warning("olx-badung | listing page failed in browser: %s", e)
        return links

    load_more_clicks = 0
    max_clicks = max(1, max_pages)
    OLX_LOAD_MORE_WAIT_SEC = 4
    OLX_BUTTON_WAIT_MS = 10_000  # 10s for button to stay visible when we click

    while load_more_clicks < max_clicks:
        html = page.content()
        batch = _olx_extract_item_links_from_html(html)
        for u in batch:
            if u not in seen:
                seen.add(u)
                links.append(u)

        # Scroll to bottom so "Muat lainnya" stays in view
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1.5)
        except Exception:
            pass

        # Find and click "Muat lainnya" (only clicks when button is visible)
        clicked = False
        for selector in OLX_LOAD_MORE_SELECTORS:
            try:
                btn = page.locator(selector).first
                if btn.is_visible(timeout=OLX_BUTTON_WAIT_MS):
                    btn.click()
                    clicked = True
                    load_more_clicks += 1
                    log.info("olx-badung | clicked 'Muat lainnya' (click %s/%s), total links so far: %s", load_more_clicks, max_clicks, len(links))
                    time.sleep(OLX_LOAD_MORE_WAIT_SEC)
                    if load_more_clicks >= max_clicks:
                        break
                    break
            except Exception:
                continue
        if not clicked or load_more_clicks >= max_clicks:
            if not clicked:
                log.info("olx-badung | 'Muat lainnya' not visible after %s clicks (page may be fully loaded)", load_more_clicks)
            break

    log.info("olx-badung | collected %s unique property URLs (%s load-more clicks)", len(links), load_more_clicks)
    return links


# Timeout for a single listing detail page load; if exceeded, skip to next listing.
DETAIL_PAGE_TIMEOUT_MS = 30_000


def olx_parse_detail_playwright(page, url: str, delay: float = 0.0) -> dict | None:
    """
    Fetch and parse a single OLX detail page into our raw row schema, using the
    already-open visible Chromium page.

    We only rely on very generic HTML patterns (h1 title, Rp price text, a
    location/breadcrumb area, description and first image) to keep it robust
    against minor layout changes. If the page does not finish loading within
    DETAIL_PAGE_TIMEOUT_MS (30s), we skip this listing and return None.
    """
    try:
        page.set_extra_http_headers(_olx_headers())
        log.info("olx-badung | detail in browser %s", url)
        # Wait for the full page load; 30s timeout then move to next listing.
        page.goto(url, wait_until="load", timeout=DETAIL_PAGE_TIMEOUT_MS)
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            # If network never fully goes idle, we still proceed with whatever DOM we have.
            pass
        try:
            page.wait_for_selector("h1", timeout=10_000)
        except Exception:
            pass
        try:
            page.evaluate("window.scrollBy(0, 100)")
        except Exception:
            pass
    except Exception as e:
        err_msg = str(e).lower()
        if "timeout" in err_msg or "timed out" in err_msg:
            log.warning("olx-badung | detail page timeout (30s), skipping: %s", url)
        else:
            log.warning("olx-badung | detail failed in browser %s: %s", url, e)
        return None

    # ------------------------------------------------------------------
    # Primary data source: structured JSON from window.__APP, via JS
    # evaluation (avoids brittle regex over HTML/inline scripts).
    # ------------------------------------------------------------------
    title = ""
    price_text = ""
    location = ""
    description = ""
    bedrooms = ""
    bathrooms = ""
    land_size_m2 = ""
    building_size_m2 = ""
    certificate = ""
    facilities_list: list[str] = []
    agent_name = ""
    agent_phone = ""
    updated_by = ""
    main_image = ""
    image_urls: list[str] = []

    app_obj = None
    try:
        # For now we ignore the clickable phone number and just focus on the
        # structured listing data available in window.__APP. The phone click /
        # Akamai timing logic can be re-enabled later if needed.
        app_obj = page.evaluate("() => (window.__APP || null)")
    except Exception:
        app_obj = None

    if isinstance(app_obj, dict):
        try:
            items_state = (app_obj.get("states") or {}).get("items") or {}
            elements = items_state.get("elements") or {}
            ad_obj = None
            if isinstance(elements, dict) and elements:
                ad_obj = next(iter(elements.values()))
            if isinstance(ad_obj, dict):
                # Title
                title = (ad_obj.get("title") or ad_obj.get("subject") or title or "").strip()

                # Price → normalized "Rp ..." string
                price = ad_obj.get("price") or {}
                if isinstance(price, dict):
                    raw_val = None
                    value = price.get("value") or {}
                    if isinstance(value, dict):
                        raw_val = value.get("raw") or value.get("value")
                    if raw_val is None:
                        raw_val = price.get("raw") or price.get("value")
                    try:
                        amount = int(raw_val)
                        price_text = f"Rp {amount:,}".replace(",", ".")
                    except Exception:
                        pass
                if not price_text and isinstance(price, dict):
                    # If OLX already has a nice "Rp ..." display string, prefer that.
                    display = (price.get("display") or "").strip()
                    if display.startswith("Rp"):
                        price_text = display

                # Location (simple human label)
                loc_obj = ad_obj.get("location") or ad_obj.get("locations_resolved") or {}
                if isinstance(loc_obj, dict):
                    location = (
                        loc_obj.get("label")
                        or loc_obj.get("display_name")
                        or loc_obj.get("city_name")
                        or location
                    ) or ""
                    location = (location or "").strip()

                # Description
                description = (ad_obj.get("description") or description or "").strip()

                # Main info string, e.g. "3 KT - 2 KM - 250 m2"
                main_info = (ad_obj.get("main_info") or "").strip()
                if main_info:
                    m_bed = re.search(r"(\d+)\s*KT", main_info, re.I)
                    if m_bed:
                        bedrooms = m_bed.group(1)
                    m_bath = re.search(r"(\d+)\s*KM", main_info, re.I)
                    if m_bath:
                        bathrooms = m_bath.group(1)
                    m_bld = list(re.finditer(r"(\d+)\s*m2", main_info, re.I))
                    if m_bld:
                        building_size_m2 = m_bld[-1].group(1)

                # Parameters (land size, certificate, facilities, and possibly beds/baths)
                params = ad_obj.get("parameters") or []
                if isinstance(params, list):
                    for p in params:
                        if not isinstance(p, dict):
                            continue
                        key = (p.get("key") or "").strip()
                        val = (
                            p.get("formatted_value")
                            or p.get("value_text")
                            or p.get("string_value")
                            or p.get("value")
                            or ""
                        )
                        val = str(val).strip()
                        if not key or not val:
                            continue
                        if key == "p_sqr_land" and not land_size_m2:
                            m_land = re.search(r"(\d+)", val)
                            if m_land:
                                land_size_m2 = m_land.group(1)
                        elif key == "p_certificate" and not certificate:
                            certificate = val
                        elif key == "p_facility":
                            facilities_list.append(val)
                        elif key in ("p_bedroom", "p_bedrooms") and not bedrooms:
                            m_bed = re.search(r"(\d+)", val)
                            if m_bed:
                                bedrooms = m_bed.group(1)
                        elif key in ("p_bathroom", "p_bathrooms") and not bathrooms:
                            m_bath = re.search(r"(\d+)", val)
                            if m_bath:
                                bathrooms = m_bath.group(1)

                # Agent / user name
                agent_name = (
                    ad_obj.get("user_name")
                    or (ad_obj.get("user") or {}).get("name")
                    or agent_name
                    or ""
                ).strip()

                # If still missing, fall back to the global users state using
                # the ad's user_id. This is where names like "Destiny Realty"
                # live for OLX property agents.
                if not agent_name:
                    user_id = str(ad_obj.get("user_id") or "").strip()
                    users_state = (app_obj.get("states") or {}).get("users") or {}
                    users_elements = users_state.get("elements") or {}
                    if user_id and isinstance(users_elements, dict):
                        user_obj = users_elements.get(user_id) or {}
                        if isinstance(user_obj, dict):
                            agent_name = (user_obj.get("name") or "").strip() or agent_name

                # Updated-by metadata (best-effort)
                updated_by = (ad_obj.get("display_date") or updated_by or "").strip()

                # Images
                imgs = ad_obj.get("images") or []
                if isinstance(imgs, list):
                    for img in imgs:
                        if not isinstance(img, dict):
                            continue
                        src = (
                            img.get("url")
                            or img.get("big")
                            or img.get("large")
                            or img.get("small")
                            or ""
                        )
                        src = str(src).strip()
                        if not src or not src.startswith("http"):
                            continue
                        if src not in image_urls:
                            image_urls.append(src)
                    if image_urls and not main_image:
                        main_image = image_urls[0]
        except Exception:
            # If JSON parsing fails for any reason, we'll fall back to HTML below.
            pass

    # ------------------------------------------------------------------
    # HTML-based fallbacks for any missing fields
    # ------------------------------------------------------------------
    html = page.content()

    # In some anti-bot / edge-cache failure cases OLX serves a raw CSS
    # asset (e.g. desktop-7904.olx.css) instead of the expected HTML
    # document for the detail URL. When that happens, page.content() is
    # just a long CSS blob and our parser sees no <h1> / "Rp ..." price,
    # which leads to "no data collected" even though the visible browser
    # might later navigate elsewhere.
    #
    # Heuristic: if the DOM snapshot we are about to parse looks non-HTML
    # (no <html> or <body> tag at all), retry once with a full "load"
    # wait. If we still don't see an HTML document, log and give up on
    # this URL so we don't keep logging huge CSS blobs.
    lower_html = (html or "").lower()
    if "<html" not in lower_html and "<body" not in lower_html:
        log.warning(
            "olx-badung | detail %s content looks non-HTML before parsing "
            "(len=%s); retrying once with wait_until=load",
            url,
            len(html or ""),
        )
        try:
            page.goto(url, wait_until="load", timeout=DETAIL_PAGE_TIMEOUT_MS)
            html = page.content()
            lower_html = (html or "").lower()
        except Exception as e:
            log.warning(
                "olx-badung | retry detail load failed for %s: %s", url, e
            )
            return None

        if "<html" not in lower_html and "<body" not in lower_html:
            log.warning(
                "olx-badung | detail %s still non-HTML after retry; skipping "
                "this URL (likely CSS/security asset instead of listing)",
                url,
            )
            return None

    soup = BeautifulSoup(html, "html.parser")

    if not title:
        h1 = soup.find("h1")
        if h1:
            title = h1.get_text(strip=True).replace(",", "")

    if not price_text:
        price_node = soup.find(string=re.compile(r"Rp\s*[\d\.]+", re.I))
        if price_node:
            price_text = price_node.strip()
    # Ensure price_text looks like a real "Rp ..." price; if JSON extraction
    # accidentally captured non-price content (e.g. a CSS blob), re-scan the
    # HTML and fall back to a sane value or empty.
    if price_text and not re.search(r"Rp\s*[\d\.]+", price_text, re.I):
        price_node = soup.find(string=re.compile(r"Rp\s*[\d\.]+", re.I))
        if price_node:
            price_text = price_node.strip()
        else:
            price_text = ""

    if not location:
        loc_candidates = []
        for el in soup.find_all(string=re.compile(r"Badung|Bali", re.I)):
            txt = el.strip()
            if 3 <= len(txt) <= 80:
                loc_candidates.append(txt)
        if loc_candidates:
            location = loc_candidates[0]

    if not description:
        d_label = soup.find(string=re.compile(r"Deskripsi", re.I))
        if d_label:
            parent = d_label.parent
            while parent and getattr(parent, "name", None) not in ("section", "div"):
                parent = parent.parent
            if parent:
                desc_block = parent.find("p") or parent.find("div")
                if desc_block:
                    description = desc_block.get_text(separator=" ", strip=True)
        if not description:
            longest = ""
            for p in soup.find_all("p"):
                txt = p.get_text(separator=" ", strip=True)
                if len(txt) > len(longest):
                    longest = txt
            description = longest

    if not main_image:
        for img in soup.find_all("img", src=True):
            src = (img.get("src") or "").strip()
            if not src:
                continue
            if "olxcdn" in src or "images.olx" in src or src.startswith("http"):
                main_image = src
                break
    if main_image and not image_urls:
        image_urls = [main_image]

    if not title and not price_text:
        # Not a valid listing page (maybe removed). Log a short snippet of the
        # page so we can see what OLX actually returned (e.g. CSS blob,
        # security wall, etc.) when debugging "no data collected".
        try:
            raw_html = page.content()
            snippet = (raw_html or "")[:1000]
            log.warning(
                "olx-badung | empty title+price for %s; first 1000 chars of page:\n%s",
                url,
                snippet,
            )
        except Exception:
            pass
        return None

    # Merge phone into agent_name so it ends up in the canonical CSV column
    # without changing the shared schema. Format: "Name | PHONE".
    if agent_phone:
        if agent_name:
            agent_name = f"{agent_name} | {agent_phone}"
        else:
            agent_name = agent_phone

    row = {
        "title": title,
        "price": price_text,
        "location": location,
        "description": description,
        "main_image": main_image,
        "image_urls": " | ".join(image_urls) if image_urls else "",
        # Rumah123-style spec keys so flatten_row(source="rumah123") can populate canonical columns
        "Kamar_Tidur": bedrooms,
        "Kamar_Mandi": bathrooms,
        "Luas_Tanah": land_size_m2,
        "Luas_Bangunan": building_size_m2,
        "Sertifikat": certificate,
        # Facilities and agent/update metadata
        "facilities_rumah": " | ".join(dict.fromkeys(facilities_list)) if facilities_list else "",
        "agent_name": agent_name,
        "updated_by": updated_by,
        "url": url,
    }
    return row


def run_olx_badung(
    args,
    csv_handle: tuple | None = None,
    existing_urls: set[str] | None = None,
) -> list[dict] | list[str]:
    """
    Scrape OLX Badung category and emit rows in the unified CSV schema.

    - With --list-only: return a list of URLs.
    - Otherwise: write rows to CSV (if csv_handle set) and return flattened dicts.
    """
    log.info("olx-badung | starting scrape (max 'Muat lainnya' clicks=%s) using visible Chromium", args.pages)
    max_pages = max(1, args.pages)  # for OLX: number of Load-more button clicks, not URL pages

    login_wait = not getattr(args, "olx_skip_login_wait", True)
    page, browser, pw = olx_get_browser(login_wait=login_wait)
    try:
        links = olx_fetch_listing_links_playwright(page, max_pages=max_pages)
        if not links:
            log.warning("olx-badung | no listing links found")
            return [] if not args.list_only else []

        if args.list_only:
            # Just return unique URLs so caller can write .txt
            return list(dict.fromkeys(links))

        seen = existing_urls if existing_urls is not None else set()
        links_to_fetch = [u for u in links if _normalize_listing_url(u) not in seen]
        skipped = len(links) - len(links_to_fetch)
        if skipped:
            log.info("olx-badung | skipping %s URLs already in output (resume/duplicate detection)", skipped)
        if not links_to_fetch:
            log.info("olx-badung | no new URLs to fetch")
            return []

        delay = getattr(args, "delay", 3.0)
        if getattr(args, "fast", False):
            delay = max(1.0, delay * 0.6)

        # From here on we only open detail pages; block CSS/images/fonts to save bandwidth.
        try:
            page.context.route(
                "**/*",
                lambda route, request: route.abort()
                if request.resource_type in ("image", "stylesheet", "font")
                else route.continue_(),
            )
            log.info("olx-badung | detail pages: blocking CSS/images/fonts to speed up parsing")
        except Exception as e:
            log.debug("olx-badung | could not set resource block for detail pages: %s", e)

        # Playwright sync API is not thread-safe: all page/browser calls must run on the
        # same thread that created them. So OLX detail fetch is always sequential (1 worker).
        rows: list[dict] = []
        total = len(links_to_fetch)
        for i, url in enumerate(links_to_fetch, 1):
            log.info("olx-badung | detail %s/%s %s", i, total, url)
            row = olx_parse_detail_playwright(page, url, delay=delay)
            if not row:
                log.debug("olx-badung | skip (parser returned None) %s", url)
                continue
            log.debug(
                "olx-badung | parsed row title=%r price=%r url=%s",
                row.get("title"),
                row.get("price"),
                url,
            )
            if _row_has_title_and_price(row, "rumah123"):
                flat = flatten_row(row, source="rumah123")
                if _write_row_to_csv(csv_handle, flat, seen):
                    rows.append(flat)
                    log.debug("olx-badung | wrote row for %s", url)
            else:
                log.warning(
                    "olx-badung | row missing title or parsable price, skipping: title=%r price=%r url=%s",
                    row.get("title"),
                    row.get("price"),
                    url,
                )

        return rows
    finally:
        try:
            browser.close()
            pw.stop()
        except Exception:
            pass


def run_villa_bali(
    args,
    csv_handle: tuple | None = None,
    existing_urls: set[str] | None = None,
) -> list[dict] | list[str] | None:
    log.info("villa-bali | launching browser (may hit Cloudflare)")
    seen = existing_urls if existing_urls is not None else set()
    page, browser, pw = villa_bali_get_browser(headless=args.headless)
    try:
        links = villa_bali_fetch_listing_links(page)
        log.info("villa-bali | collected %s villa URLs", len(links))
        if not links:
            log.warning("villa-bali | no listing links found")
            return []
        if args.list_only:
            return links  # caller writes .txt
        links_to_fetch = [u for u in links if _normalize_listing_url(u) not in seen]
        skipped = len(links) - len(links_to_fetch)
        if skipped:
            log.info("villa-bali | skipping %s URLs already in output (resume/duplicate detection)", skipped)
        if not links_to_fetch:
            log.info("villa-bali | no new URLs to fetch")
            return []
        workers = max(1, getattr(args, "workers", 3))
        workers = min(workers, len(links_to_fetch))
        rows = []
        if workers <= 1:
            log.info("villa-bali | fetching %s detail pages (delay %ss)", len(links_to_fetch), args.delay)
            for i, url in enumerate(links_to_fetch, 1):
                log.debug("villa-bali | [%s/%s] %s", i, len(links_to_fetch), url)
                row = villa_bali_parse_detail(page, url, delay=args.delay)
                if row and _row_has_title_and_price(row, "villa-bali"):
                    flat = flatten_row(row, source="villa-bali")
                    if _write_row_to_csv(csv_handle, flat, seen):
                        rows.append(flat)
                        log.debug("villa-bali | ok [%s/%s] %s", i, len(links_to_fetch), url)
                else:
                    log.debug("villa-bali | skip (no title/price) [%s/%s] %s", i, len(links_to_fetch), url)
        else:
            pages_list = [page]
            try:
                context = page.context
                for _ in range(workers - 1):
                    pages_list.append(context.new_page())
            except Exception as e:
                log.warning("villa-bali | could not create extra pages, falling back to 1 worker: %s", e)
                pages_list = [page]
                workers = 1
            if workers > 1:
                log.info("villa-bali | fetching %s detail pages with %s workers (delay %ss)", len(links_to_fetch), workers, args.delay)
                url_queue: queue.Queue[str | None] = queue.Queue()
                result_queue: queue.Queue[dict] = queue.Queue()
                for u in links_to_fetch:
                    url_queue.put(u)
                for _ in range(workers):
                    url_queue.put(None)
                total = len(links_to_fetch)
                done_count = [0]

                def villa_bali_worker(worker_page, worker_id: int) -> None:
                    while True:
                        u = url_queue.get()
                        if u is None:
                            return
                        with _csv_write_lock:
                            done_count[0] += 1
                            n = done_count[0]
                        if n % 10 == 0 or n == total:
                            log.info("villa-bali | detail %s/%s (worker %s)", n, total, worker_id)
                        try:
                            row = villa_bali_parse_detail(worker_page, u, delay=args.delay)
                        except Exception as e:
                            log.warning("villa-bali | worker %s failed for %s: %s", worker_id, u, e)
                            continue
                        if row and _row_has_title_and_price(row, "villa-bali"):
                            flat = flatten_row(row, source="villa-bali")
                            if _write_row_to_csv(csv_handle, flat, seen):
                                result_queue.put(flat)
                                log.debug("villa-bali | ok %s", u)
                        else:
                            log.debug("villa-bali | skip (no title/price) %s", u)

                threads = [
                    threading.Thread(target=villa_bali_worker, args=(pages_list[i], i + 1))
                    for i in range(workers)
                ]
                for t in threads:
                    t.start()
                for t in threads:
                    t.join()
                while True:
                    try:
                        rows.append(result_queue.get_nowait())
                    except queue.Empty:
                        break
                for i in range(1, len(pages_list)):
                    try:
                        pages_list[i].close()
                    except Exception:
                        pass
                log.info("villa-bali | parallel detail fetch done: %s rows", len(rows))
            else:
                for i, url in enumerate(links_to_fetch, 1):
                    row = villa_bali_parse_detail(page, url, delay=args.delay)
                    if row and _row_has_title_and_price(row, "villa-bali"):
                        flat = flatten_row(row, source="villa-bali")
                        if _write_row_to_csv(csv_handle, flat, seen):
                            rows.append(flat)
        return rows
    finally:
        browser.close()
        pw.stop()


def main():
    ap = argparse.ArgumentParser(
        description="Unified villa scraper: all sources → one CSV (same columns, same data strategy)"
    )
    ap.add_argument(
        "--source", "-s",
        choices=["all", "rumah123", "villa-bali", "balilongterm", "balicoconut", "balihomeimmo", "olx-badung", "balirealty"],
        default="all",
        help="Source(s) to scrape (default: all)",
    )
    ap.add_argument("--pages", type=int, default=999, help="Max listing pages per source (default 999 = no practical limit)")
    ap.add_argument("--delay", type=float, default=3.0, help="Delay between detail requests (seconds)")
    ap.add_argument("--output", "-o", default="villa_listings.csv", help="Output CSV path")
    ap.add_argument("--append", action="store_true", help="Append to existing CSV")
    ap.add_argument("--list-only", action="store_true", help="Only collect URLs to .txt (single source only)")
    ap.add_argument("--json", action="store_true", help="Also write JSON (Rumah123 only)")
    ap.add_argument("--workers", "-w", type=int, default=3, help="Number of concurrent detail-page sessions (default 3). Uses multiple browser pages for OLX/Villa-Bali and parallel requests for Rumah123 fallback.")
    ap.add_argument("--no-headless", action="store_true", help="Run browser visible (default is headless); use if security verification needs to be completed manually")
    ap.add_argument("--fast", "-f", action="store_true", help="Shorter delays for all sources (faster, minimal block risk)")
    ap.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"], help="Logging level (default: INFO)")
    ap.add_argument("--log-file", metavar="PATH", default=None, help="Also write logs to this file (utf-8)")
    ap.add_argument("--resume", action="store_true", default=True, help="Load existing URLs from output file and skip duplicates (default: on)")
    ap.add_argument("--no-resume", action="store_false", dest="resume", help="Disable resume; do not load existing URLs (fetch and write all)")
    ap.add_argument(
        "--no-proxy", "-n",
        action="store_true",
        dest="no_proxy",
        help="Disable proxy rotation and run all network calls directly (proxyless mode)",
    )
    ap.add_argument(
        "--olx-skip-login-wait",
        action="store_true",
        default=True,
        help="For OLX only: do not wait on the home page before scraping (default: no wait).",
    )
    ap.add_argument(
        "--olx-login-wait",
        action="store_false",
        dest="olx_skip_login_wait",
        help="For OLX only: wait 60s on the home page so you can complete login/verification manually.",
    )
    args = ap.parse_args()
    args.headless = not getattr(args, "no_headless", False)  # default: headless True

    # Apply global proxyless mode if requested from CLI.
    if getattr(args, "no_proxy", False):
        _switch_to_proxyless()

    setup_logging(level=args.log_level, log_file=args.log_file)
    log.info("source=%s output=%s append=%s headless=%s", args.source, args.output, args.append, args.headless)

    # "all" = Rumah123, BLT, BCL, Bali Home Immo (yearly+monthly) + OLX Badung
    sources_to_run = (
        ["rumah123", "balilongterm", "balicoconut", "balihomeimmo", "olx-badung", "balirealty"]
        if args.source == "all"
        else [args.source]
    )

    if args.source == "all" and args.list_only:
        log.error("--list-only is not supported with --source all; pick a single source")
        return

    all_rows = []
    out_path = Path(args.output)
    file_exists = out_path.exists()
    # With resume (default), never truncate: append if file exists so previous session data is kept
    use_append = args.append or (getattr(args, "resume", True) and file_exists)
    write_header = not (use_append and file_exists)

    # Resume & duplicate detection (default: on). Load existing URLs from output CSV/Excel; skip those when fetching and never write the same listing twice.
    if getattr(args, "resume", True):
        existing_urls = _load_existing_urls(out_path)
        if existing_urls:
            log.info("Resume: loaded %s existing URLs from %s (CSV rows). Will skip those when fetching and prevent double entries.", len(existing_urls), out_path)
        else:
            log.info("Resume: no existing URLs in %s; starting fresh (duplicate prevention still on).", out_path)
    else:
        existing_urls = set()
        log.info("Resume disabled (--no-resume); will not skip existing URLs.")

    if args.source == "all":
        # Run sources sequentially to limit resource and bandwidth use
        workers_to_run = [
            s for s in sources_to_run
            if (s != "balilongterm" or run_balilongterm)
            and (s != "balicoconut" or run_balicoconut)
            and (s != "balihomeimmo" or run_balihomeimmo)
            and (s != "balirealty" or run_balirealty)
        ]
        if "balihomeimmo" in sources_to_run and run_balihomeimmo is None:
            log.warning("skipping balihomeimmo (scrape_balihomeimmo not importable)")
        for src in workers_to_run:
            log.info("--- %s ---", src)
            try:
                if src == "rumah123":
                    result = run_rumah123(args, existing_urls=existing_urls)
                elif src == "villa-bali":
                    result = run_villa_bali(args, existing_urls=existing_urls)
                elif src == "olx-badung":
                    result = run_olx_badung(args, existing_urls=existing_urls)
                elif src == "balilongterm":
                    result = run_balilongterm(args)
                elif src == "balicoconut":
                    result = run_balicoconut(args)
                elif src == "balihomeimmo":
                    result = run_balihomeimmo(args)
                elif src == "balirealty":
                    result = run_balirealty(args)
                else:
                    result = None
            except Exception as e:
                log.warning("%s failed: %s", src, e)
                continue
            if result is None:
                log.warning("%s returned no data", src)
                continue
            if isinstance(result, list) and result and isinstance(result[0], str):
                out_path = Path(args.output).with_suffix(".txt")
                out_path.write_text("\n".join(result), encoding="utf-8")
                log.info("wrote %s URLs to %s", len(result), out_path)
                return
            rows = result if isinstance(result, list) else []
            if rows:
                # Skip duplicates: only write rows whose url (normalized) is not already in existing_urls
                new_rows = []
                for r in rows:
                    url = (r.get("url") or "").strip()
                    key = _normalize_listing_url(url) if url else ""
                    if key and key not in existing_urls:
                        new_rows.append(r)
                        existing_urls.add(key)
                if new_rows:
                    mode = "a" if use_append else "w"
                    with open(out_path, mode, newline="", encoding="utf-8") as f:
                        w = csv.DictWriter(f, fieldnames=MAIN_CSV_COLUMNS, extrasaction="ignore")
                        if write_header:
                            w.writeheader()
                            write_header = False
                        for r in new_rows:
                            safe = {k: _csv_cell(str(r.get(k, ""))) for k in MAIN_CSV_COLUMNS}
                            w.writerow(safe)
                    log.info("→ %s: wrote %s new rows (skipped %s duplicates) to %s", src, len(new_rows), len(rows) - len(new_rows), out_path)
                else:
                    log.info("→ %s: all %s rows already in file (duplicates skipped)", src, len(rows))
                all_rows.extend(rows)
            else:
                log.warning("%s returned no data", src)
    else:
        # Single source: open CSV once (append when resuming so we never erase previous data)
        with open(out_path, "a" if use_append else "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=MAIN_CSV_COLUMNS, extrasaction="ignore")
            if write_header:
                w.writeheader()
                f.flush()
            csv_handle = (w, f)
            for src in sources_to_run:
                log.info("--- %s ---", src)
                if src == "rumah123":
                    result = run_rumah123(args, csv_handle=csv_handle, existing_urls=existing_urls)
                elif src == "villa-bali":
                    result = run_villa_bali(args, csv_handle=csv_handle, existing_urls=existing_urls)
                elif src == "olx-badung":
                    result = run_olx_badung(args, csv_handle=csv_handle, existing_urls=existing_urls)
                elif src == "balilongterm":
                    if run_balilongterm is None:
                        log.warning("skipping balilongterm (scrape_balilongterm not importable)")
                        continue
                    result = run_balilongterm(args)
                elif src == "balicoconut":
                    if run_balicoconut is None:
                        log.warning("skipping balicoconut (scrape_balicoconut not importable)")
                        continue
                    result = run_balicoconut(args)
                elif src == "balihomeimmo":
                    if run_balihomeimmo is None:
                        log.warning("skipping balihomeimmo (scrape_balihomeimmo not importable)")
                        continue
                    result = run_balihomeimmo(args)
                elif src == "balirealty":
                    if run_balirealty is None:
                        log.warning("skipping balirealty (scrape_balirealty not importable)")
                        continue
                    result = run_balirealty(args, csv_handle=csv_handle, existing_urls=existing_urls)
                else:
                    continue

                if result is None:
                    log.warning("%s returned no data", src)
                    continue
                if isinstance(result, list) and result and isinstance(result[0], str):
                    out_path = Path(args.output).with_suffix(".txt")
                    out_path.write_text("\n".join(result), encoding="utf-8")
                    log.info("wrote %s URLs to %s", len(result), out_path)
                    return
                rows = result if isinstance(result, list) else []
                if rows:
                    all_rows.extend(rows)
                    # BLT/BCL/BHI don't support csv_handle; write only new rows (skip duplicates)
                    for r in rows:
                        url = (r.get("url") or "").strip()
                        key = _normalize_listing_url(url) if url else ""
                        if key and key in existing_urls:
                            continue
                        safe = {k: _csv_cell(str(r.get(k, ""))) for k in MAIN_CSV_COLUMNS}
                        w.writerow(safe)
                        f.flush()
                        if key:
                            existing_urls.add(key)

    if args.source != "all":
        if not all_rows:
            log.warning("no data collected")
            return
        # Rows were already written incrementally to CSV; no bulk write needed
        mode = "appended" if args.append else "wrote"
        log.info("%s %s rows to %s", mode.capitalize(), len(all_rows), out_path)

        if args.json and args.source == "rumah123" and all_rows:
            # Write normalized portal-style JSON structure:
            # {"villas": [ { ...canonical villa fields... }, ... ]}
            villas_payload = [_row_to_portal_villa(row, source="rumah123") for row in all_rows]
            json_path = out_path.with_suffix(".json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump({"villas": villas_payload}, f, ensure_ascii=False, indent=2)
            log.info("wrote %s in portal JSON format (villas[])", json_path)
    else:
        log.info("all sources done | total rows in %s: %s", out_path, len(all_rows))


if __name__ == "__main__":
    main()
